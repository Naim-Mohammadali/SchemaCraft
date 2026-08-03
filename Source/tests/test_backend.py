from __future__ import annotations

import base64
import copy
import json
import tempfile
import threading
import time
import unittest
import zipfile
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path
from unittest import mock
from urllib.error import HTTPError
from urllib.request import Request, urlopen

from openpyxl import load_workbook

import SchemaCraft as APP


CAT_IDENTITY = "cat_000000000001"
CAT_DOCUMENTS = "cat_000000000002"
CAT_CHILDREN = "cat_000000000003"

FLD_NAME = "fld_000000000001"
FLD_FATHER = "fld_000000000002"
FLD_FAMILY = "fld_000000000003"
FLD_WORKS = "fld_000000000004"
FLD_GREGORIAN = "fld_000000000005"
FLD_HIJRI = "fld_000000000006"
FLD_SHAMSI = "fld_000000000007"
FLD_NUMBER = "fld_000000000008"
FLD_STATUS = "fld_000000000009"
FLD_NOTES = "fld_00000000000a"
FLD_DOC_TYPE = "fld_00000000000b"
FLD_DOC_DATE = "fld_00000000000c"
FLD_CUSTOM_FILE = "fld_00000000000d"
FLD_ORIGINAL_FILE = "fld_00000000000e"
FLD_CHILD_NAME = "fld_00000000000f"
FLD_CHILD_DATE = "fld_000000000010"
FLD_VERIFIED = "fld_000000000011"
FLD_SKILLS = "fld_000000000012"

COND_DOCUMENTS = "cond_000000000001"
COND_NOTES = "cond_000000000002"


def field(
    field_id: str,
    label: str,
    field_type: str,
    *,
    required: bool = False,
    options: list[str] | None = None,
    searchable: bool = False,
    search_match: str | None = None,
    show: bool = False,
    title: bool = False,
    naming: dict | None = None,
) -> dict:
    result = {
        "id": field_id,
        "label": label,
        "type": field_type,
        "required": required,
        "placeholder": "",
        "width": "normal",
        "options": options or [],
        "searchable": searchable,
        "search_match": search_match
        or ("contains" if field_type in {"text", "textarea"} else "exact"),
        "show_in_results": show,
        "result_title": title,
    }
    if field_type == "file":
        result["file_naming"] = naming or {"mode": "original", "parts": []}
    return result


def configured_schema(revision: int = 0) -> dict:
    return {
        "schema_version": 1,
        "revision": revision,
        "app": {
            "title": "نظام تجريبي",
            "entity_singular": "شخص",
            "entity_plural": "الأشخاص",
            "direction": "rtl",
            "primary_color": "#315F8A",
            "background_color": "#F5F7FA",
            "surface_color": "#FFFFFF",
        },
        "categories": [
            {
                "id": CAT_IDENTITY,
                "label": "البيانات الأساسية",
                "description": "بيانات الشخص الرئيسية.",
                "kind": "main",
                "add_label": "",
                "auto_start": False,
                "anchor_field_id": None,
                "fields": [
                    field(
                        FLD_NAME,
                        "الاسم",
                        "text",
                        searchable=True,
                        show=True,
                        title=True,
                    ),
                    field(
                        FLD_FATHER,
                        "اسم الأب",
                        "text",
                        searchable=True,
                        show=True,
                        title=True,
                    ),
                    field(
                        FLD_FAMILY,
                        "الشهرة",
                        "text",
                        searchable=True,
                        show=True,
                        title=True,
                    ),
                    field(
                        FLD_WORKS,
                        "هل يعمل",
                        "yes_no",
                        options=["نعم", "لا"],
                        searchable=True,
                        show=True,
                    ),
                    field(
                        FLD_GREGORIAN,
                        "تاريخ ميلادي",
                        "date_gregorian",
                        searchable=True,
                    ),
                    field(FLD_HIJRI, "تاريخ هجري", "date_hijri"),
                    field(FLD_SHAMSI, "تاريخ هجري شمسي", "date_persian"),
                    field(FLD_NUMBER, "رقم", "number"),
                    field(
                        FLD_STATUS,
                        "الحالة",
                        "select",
                        options=["نشط", "متوقف"],
                        searchable=True,
                    ),
                    field(FLD_NOTES, "ملاحظات", "textarea"),
                    field(
                        FLD_VERIFIED,
                        "تم التحقق",
                        "checkbox",
                        searchable=True,
                    ),
                    field(
                        FLD_SKILLS,
                        "المهارات",
                        "checkbox_group",
                        options=["برمجة", "إدارة"],
                        searchable=True,
                    ),
                ],
            },
            {
                "id": CAT_DOCUMENTS,
                "label": "المرفقات",
                "description": "ملفات الشخص.",
                "kind": "repeatable",
                "add_label": "إضافة مرفق",
                "auto_start": False,
                "anchor_field_id": FLD_WORKS,
                "fields": [
                    field(
                        FLD_DOC_TYPE,
                        "نوع المستند",
                        "text",
                        required=True,
                        searchable=True,
                        show=True,
                    ),
                    field(
                        FLD_DOC_DATE,
                        "تاريخ المستند",
                        "date_gregorian",
                        searchable=True,
                    ),
                    field(
                        FLD_CUSTOM_FILE,
                        "الملف المسمّى",
                        "file",
                        naming={
                            "mode": "template",
                            "parts": [
                                {"field_id": FLD_NAME, "prefix": ""},
                                {"field_id": FLD_FATHER, "prefix": " "},
                                {"field_id": FLD_FAMILY, "prefix": " "},
                                {"field_id": FLD_DOC_TYPE, "prefix": " - "},
                            ],
                        },
                    ),
                    field(
                        FLD_ORIGINAL_FILE,
                        "الملف باسمه الأصلي",
                        "file",
                        naming={"mode": "original", "parts": []},
                    ),
                ],
            },
            {
                "id": CAT_CHILDREN,
                "label": "الأولاد",
                "description": "",
                "kind": "repeatable",
                "add_label": "إضافة ولد",
                "auto_start": True,
                "anchor_field_id": None,
                "fields": [
                    field(
                        FLD_CHILD_NAME,
                        "اسم الولد",
                        "text",
                        searchable=True,
                        show=True,
                    ),
                    field(
                        FLD_CHILD_DATE,
                        "تاريخ الولادة",
                        "date_hijri",
                        searchable=True,
                    ),
                ],
            },
        ],
        "conditions": [
            {
                "id": COND_DOCUMENTS,
                "target_type": "category",
                "target_id": CAT_DOCUMENTS,
                "source_field_id": FLD_WORKS,
                "operator": "equals",
                "value": "نعم",
            },
            {
                "id": COND_NOTES,
                "target_type": "field",
                "target_id": FLD_NOTES,
                "source_field_id": FLD_STATUS,
                "operator": "equals",
                "value": "نشط",
            },
        ],
    }


def valid_code(number: int) -> str:
    return f"T{number:07d}"


class GenericDataEntryBackendTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temporary = tempfile.TemporaryDirectory()
        self.root = Path(self.temporary.name)
        self.originals = {
            "DATA_DIR": APP.DATA_DIR,
            "SCHEMA_PATH": APP.SCHEMA_PATH,
            "WORKBOOK_PATH": APP.WORKBOOK_PATH,
            "BACKUP_DIR": APP.BACKUP_DIR,
            "STARTUP_ERROR_LOG": APP.STARTUP_ERROR_LOG,
            "BUILDER_AUTH_PATH": APP.BUILDER_AUTH_PATH,
            "DEVELOPER_MODE": APP.DEVELOPER_MODE,
            "_BUILDER_UNLOCKED": APP._BUILDER_UNLOCKED,
            "_DATASET_SNAPSHOT": APP._DATASET_SNAPSHOT,
        }
        APP.DATA_DIR = self.root
        APP.SCHEMA_PATH = self.root / "schema.json"
        APP.WORKBOOK_PATH = self.root / "database.xlsx"
        APP.BACKUP_DIR = self.root / "backups"
        APP.STARTUP_ERROR_LOG = self.root / "startup-error.log"
        APP.BUILDER_AUTH_PATH = self.root / "builder-auth.json"
        APP.DEVELOPER_MODE = True
        APP._BUILDER_UNLOCKED = False
        APP.invalidate_dataset_cache()
        APP.ensure_storage()

    def tearDown(self) -> None:
        for name, value in self.originals.items():
            setattr(APP, name, value)
        self.temporary.cleanup()

    def configure(self) -> dict:
        return APP.save_schema(configured_schema())

    def create_person(
        self,
        *,
        code: str = "A12F12B7",
        name: str = "عَلِي",
        works: str = "نعم",
        documents: list[dict] | None = None,
        children: list[dict] | None = None,
    ) -> dict:
        return APP.save_record(
            {
                "mode": "create",
                "record_code": code,
                "main": {
                    FLD_NAME: name,
                    FLD_FATHER: "حسن",
                    FLD_FAMILY: "محمدي",
                    FLD_WORKS: works,
                    FLD_GREGORIAN: "2001-04-09",
                    FLD_HIJRI: "1447-09-12",
                    FLD_SHAMSI: "1405-01-20",
                    FLD_NUMBER: "12.5",
                    FLD_STATUS: "نشط",
                    FLD_NOTES: "ملاحظة",
                    FLD_VERIFIED: True,
                    FLD_SKILLS: ["برمجة"],
                },
                "related": {
                    CAT_DOCUMENTS: documents or [],
                    CAT_CHILDREN: children
                    or [
                        {
                            "values": {
                                FLD_CHILD_NAME: "سارة",
                                FLD_CHILD_DATE: "1445-03-10",
                            }
                        }
                    ],
                },
            }
        )

    def test_01_initial_package_is_empty(self) -> None:
        schema = APP.schema_response(APP.read_schema_file())
        self.assertEqual(schema["categories"], [])
        self.assertEqual(schema["conditions"], [])
        self.assertEqual(schema["stats"], {
            "record_count": 0,
            "field_count": 0,
            "category_count": 0,
        })
        workbook = load_workbook(APP.WORKBOOK_PATH)
        try:
            self.assertEqual(workbook.sheetnames, [APP.MAIN_SHEET, APP.META_SHEET])
            self.assertTrue(workbook[APP.MAIN_SHEET].row_dimensions[1].hidden)
            self.assertEqual(workbook[APP.MAIN_SHEET].max_row, 2)
        finally:
            workbook.close()

    def test_02_schema_supports_every_field_and_category_type(self) -> None:
        saved = self.configure()
        self.assertEqual(saved["stats"]["category_count"], 3)
        self.assertEqual(saved["stats"]["field_count"], 18)
        types = {
            field["type"]
            for category in saved["categories"]
            for field in category["fields"]
        }
        # The legacy all-types fixture intentionally omits the optional,
        # read-only system metadata fields. They have dedicated regressions.
        self.assertEqual(types, APP.FIELD_TYPES - APP.SYSTEM_FIELD_TYPES)
        self.assertEqual(
            {category["kind"] for category in saved["categories"]},
            {"main", "repeatable"},
        )
        workbook = load_workbook(APP.WORKBOOK_PATH)
        try:
            self.assertEqual(len(workbook.sheetnames), 4)
            main = workbook[APP.MAIN_SHEET]
            headers = APP.technical_headers(main)
            self.assertIn(FLD_NAME, headers)
            self.assertEqual(
                main.cell(2, headers.index(FLD_NAME) + 1).value,
                "الاسم",
            )
            self.assertTrue(main.row_dimensions[1].hidden)
        finally:
            workbook.close()

    def test_03_all_three_calendars_validate_and_store_typed_values(self) -> None:
        self.configure()
        self.create_person()
        loaded = APP.load_record("A12F12B7")
        self.assertEqual(loaded["main"][FLD_GREGORIAN], "2001-04-09")
        self.assertEqual(loaded["main"][FLD_HIJRI], "1447-09-12")
        self.assertEqual(loaded["main"][FLD_SHAMSI], "1405-01-20")

        invalid_cases = [
            (FLD_GREGORIAN, "2024-02-30", "ميلادي"),
            (FLD_HIJRI, "1446-02-30", "هجري"),
            (FLD_SHAMSI, "1404-12-30", "شمسي"),
        ]
        for index, (field_id, value, message) in enumerate(invalid_cases, start=2):
            with self.subTest(field_id=field_id):
                with self.assertRaisesRegex(APP.ApplicationError, message):
                    APP.save_record(
                        {
                            "mode": "create",
                            "record_code": f"D000000{index}",
                            "main": {field_id: value},
                            "related": {},
                        }
                    )

    def test_04_blank_record_allowed_when_no_required_visible_field(self) -> None:
        self.configure()
        result = APP.save_record(
            {
                "mode": "create",
                "record_code": "B0000001",
                "main": {},
                "related": {},
            }
        )
        self.assertEqual(result["action"], "created")
        self.assertEqual(APP.load_record("B0000001")["record_code"], "B0000001")

    def test_05_conditions_minor_ids_and_hidden_data(self) -> None:
        self.configure()
        hidden = self.create_person(
            code="C0000001",
            works="لا",
            documents=[
                {
                    "values": {
                        FLD_DOC_TYPE: "يجب تجاهله",
                        FLD_DOC_DATE: "2024-01-01",
                    }
                }
            ],
        )
        self.assertEqual(hidden["related_rows"], 2)
        loaded_hidden = APP.load_record("C0000001")
        hidden_documents = loaded_hidden["related"][CAT_DOCUMENTS]
        self.assertEqual(len(hidden_documents), 1)
        self.assertEqual(
            hidden_documents[0]["values"][FLD_DOC_TYPE],
            "يجب تجاهله",
        )
        self.assertEqual(loaded_hidden["main"][FLD_NOTES], "ملاحظة")

        visible = self.create_person(
            code="C0000002",
            documents=[
                {"values": {FLD_DOC_TYPE: "جواز"}},
                {"values": {FLD_DOC_TYPE: "هوية"}},
            ],
        )
        self.assertEqual(visible["related_rows"], 3)
        rows = APP.load_record("C0000002")["related"][CAT_DOCUMENTS]
        self.assertEqual([row["minor_id"] for row in rows], [1, 2])
        self.assertEqual(len({row["_child_id"] for row in rows}), 2)

    def test_06_required_select_and_number_validation(self) -> None:
        self.configure()
        with self.assertRaisesRegex(APP.ApplicationError, "نوع المستند"):
            self.create_person(
                documents=[{"values": {FLD_DOC_DATE: "2024-01-01"}}]
            )
        with self.assertRaisesRegex(APP.ApplicationError, "الحالة"):
            APP.save_record(
                {
                    "mode": "create",
                    "record_code": "V0000001",
                    "main": {FLD_STATUS: "قيمة غير معروفة"},
                    "related": {},
                }
            )
        with self.assertRaisesRegex(APP.ApplicationError, "رقمًا"):
            APP.save_record(
                {
                    "mode": "create",
                    "record_code": "V0000002",
                    "main": {FLD_NUMBER: "abc"},
                    "related": {},
                }
            )

    def test_07_partial_arabic_and_grouped_related_search(self) -> None:
        self.configure()
        self.create_person(
            children=[
                {
                    "values": {
                        FLD_CHILD_NAME: "سارة",
                        FLD_CHILD_DATE: "1445-03-10",
                    }
                },
                {
                    "values": {
                        FLD_CHILD_NAME: "مريم",
                        FLD_CHILD_DATE: "1446-04-11",
                    }
                },
            ]
        )
        partial = APP.search_records({FLD_NAME: "علي"})
        self.assertEqual(
            [match["record_code"] for match in partial["matches"]],
            ["A12F12B7"],
        )
        grouped = APP.search_records(
            {FLD_CHILD_NAME: "سار", FLD_CHILD_DATE: "1445-03-10"}
        )
        self.assertEqual(len(grouped["matches"]), 1)
        cross_row = APP.search_records(
            {FLD_CHILD_NAME: "سار", FLD_CHILD_DATE: "1446-04-11"}
        )
        self.assertEqual(cross_row["matches"], [])
        with self.assertRaisesRegex(APP.ApplicationError, "معيار بحث"):
            APP.search_records({})

    def test_07b_temporary_search_fields_allow_non_default_fields(self) -> None:
        self.configure()
        self.create_person()
        with self.assertRaisesRegex(APP.ApplicationError, "معيار بحث"):
            APP.search_records({FLD_NOTES: "ملاحظة"})
        temporary = APP.search_records(
            {
                "_search_field_ids": [FLD_NOTES],
                FLD_NOTES: "ملاحظة",
            }
        )
        self.assertEqual(
            [match["record_code"] for match in temporary["matches"]],
            ["A12F12B7"],
        )
        with self.assertRaisesRegex(APP.ApplicationError, "غير متاح"):
            APP.search_records(
                {
                    "_search_field_ids": [FLD_CUSTOM_FILE],
                    FLD_CUSTOM_FILE: "photo.png",
                }
            )

    def test_08_update_preserves_record_and_child_identity(self) -> None:
        self.configure()
        self.create_person(
            documents=[
                {"values": {FLD_DOC_TYPE: "جواز"}},
                {"values": {FLD_DOC_TYPE: "هوية"}},
            ]
        )
        original = APP.load_record("A12F12B7")
        kept_child = original["related"][CAT_DOCUMENTS][1]
        result = APP.save_record(
            {
                "mode": "update",
                "record_code": "A12F12B7",
                "main": {
                    **original["main"],
                    FLD_NAME: "علي المعدّل",
                },
                "related": {
                    CAT_DOCUMENTS: [
                        {
                            "_child_id": kept_child["_child_id"],
                            "values": {
                                **kept_child["values"],
                                FLD_DOC_TYPE: "هوية معدلة",
                            },
                        }
                    ],
                    CAT_CHILDREN: [],
                },
            }
        )
        self.assertEqual(result["action"], "updated")
        loaded = APP.load_record("A12F12B7")
        self.assertEqual(loaded["_record_id"], original["_record_id"])
        self.assertEqual(loaded["main"][FLD_NAME], "علي المعدّل")
        self.assertEqual(len(loaded["related"][CAT_DOCUMENTS]), 1)
        self.assertEqual(
            loaded["related"][CAT_DOCUMENTS][0]["_child_id"],
            kept_child["_child_id"],
        )
        self.assertEqual(loaded["related"][CAT_DOCUMENTS][0]["minor_id"], 1)

    def test_09_schema_rename_reorder_preserves_values_by_stable_ids(self) -> None:
        self.configure()
        self.create_person()
        before = APP.load_record("A12F12B7")
        schema = APP.read_schema_file()
        identity = schema["categories"][0]
        identity["label"] = "هوية الشخص"
        identity["fields"][0]["label"] = "الاسم الكامل"
        identity["fields"].reverse()
        saved = APP.save_schema(schema)
        self.assertEqual(saved["revision"], 2)
        after = APP.load_record("A12F12B7")
        self.assertEqual(after["_record_id"], before["_record_id"])
        self.assertEqual(after["main"][FLD_NAME], "عَلِي")
        workbook = load_workbook(APP.WORKBOOK_PATH)
        try:
            headers = APP.technical_headers(workbook[APP.MAIN_SHEET])
            name_column = headers.index(FLD_NAME) + 1
            self.assertEqual(
                workbook[APP.MAIN_SHEET].cell(2, name_column).value,
                "الاسم الكامل",
            )
        finally:
            workbook.close()

        workbook = load_workbook(APP.WORKBOOK_PATH)
        try:
            main_sheet = workbook[APP.MAIN_SHEET]
            main_headers = APP.technical_headers(main_sheet)
            main_sheet.cell(
                2,
                main_headers.index(FLD_NAME) + 1,
                "الاسم المعدّل من Excel",
            )
            documents_sheet = workbook[
                APP.related_sheet_names(saved)[CAT_DOCUMENTS]
            ]
            related_headers = APP.technical_headers(documents_sheet)
            documents_sheet.cell(
                2,
                related_headers.index(FLD_DOC_TYPE) + 1,
                "نوع الملف المعدّل من Excel",
            )
            workbook.save(APP.WORKBOOK_PATH)
        finally:
            workbook.close()

        synced = APP.read_schema_with_excel_labels()
        synced_indexes = APP.schema_indexes(synced)
        self.assertEqual(synced["revision"], 3)
        self.assertEqual(
            synced_indexes["fields"][FLD_NAME]["label"],
            "الاسم المعدّل من Excel",
        )
        self.assertEqual(
            synced_indexes["fields"][FLD_DOC_TYPE]["label"],
            "نوع الملف المعدّل من Excel",
        )
        self.assertEqual(
            APP.load_record("A12F12B7")["main"][FLD_NAME],
            "عَلِي",
        )
        self.assertEqual(
            [
                match["record_code"]
                for match in APP.search_records({FLD_NAME: "علي"})["matches"]
            ],
            ["A12F12B7"],
        )

    def test_10_revision_conflict_is_rejected(self) -> None:
        schema = APP.read_schema_file()
        APP.save_schema(schema)
        with self.assertRaisesRegex(APP.ApplicationError, "تغيّر الإعداد"):
            APP.save_schema(schema)

    def test_11_attachment_template_original_counter_bytes_and_lifecycle(self) -> None:
        self.configure()
        pdf_one = b"%PDF-1.4 first"
        pdf_two = b"%PDF-1.4 second"
        image = b"\x89PNG\r\n\x1a\nimage"
        upload = lambda name, data: {
            "upload": {
                "name": name,
                "data": base64.b64encode(data).decode("ascii"),
            }
        }
        result = self.create_person(
            documents=[
                {
                    "values": {
                        FLD_DOC_TYPE: "جواز سفر",
                        FLD_CUSTOM_FILE: upload("first.pdf", pdf_one),
                        FLD_ORIGINAL_FILE: upload("photo.png", image),
                    }
                },
                {
                    "values": {
                        FLD_DOC_TYPE: "جواز سفر",
                        FLD_CUSTOM_FILE: upload("second.pdf", pdf_two),
                    }
                },
            ]
        )
        expected = {
            "attachments/عَلِي حسن محمدي - جواز سفر.pdf",
            "attachments/عَلِي حسن محمدي - جواز سفر (2).pdf",
            "attachments/photo.png",
        }
        self.assertEqual(set(result["attachment_files"]), expected)
        directory = APP.attachments_directory()
        self.assertEqual(
            (directory / "عَلِي حسن محمدي - جواز سفر.pdf").read_bytes(),
            pdf_one,
        )
        self.assertEqual(
            (directory / "عَلِي حسن محمدي - جواز سفر (2).pdf").read_bytes(),
            pdf_two,
        )
        self.assertEqual((directory / "photo.png").read_bytes(), image)

        loaded = APP.load_record("A12F12B7")
        kept = loaded["related"][CAT_DOCUMENTS][0]
        updated = APP.save_record(
            {
                "mode": "update",
                "record_code": "A12F12B7",
                "main": loaded["main"],
                "related": {
                    CAT_DOCUMENTS: [
                        {
                            "_child_id": kept["_child_id"],
                            "values": {
                                FLD_DOC_TYPE: "جواز سفر",
                                FLD_CUSTOM_FILE: {
                                    "stored_path": kept["values"][FLD_CUSTOM_FILE]
                                },
                                FLD_ORIGINAL_FILE: "",
                            },
                        }
                    ],
                    CAT_CHILDREN: [],
                },
            }
        )
        self.assertEqual(
            updated["attachment_files"],
            ["attachments/عَلِي حسن محمدي - جواز سفر.pdf"],
        )
        self.assertEqual(len(list(directory.iterdir())), 1)
        deleted = APP.delete_record("A12F12B7")
        self.assertEqual(deleted["deleted_attachment_files"], 1)
        self.assertEqual(list(directory.iterdir()), [])

    def test_12_attachment_template_requires_only_configured_parts(self) -> None:
        self.configure()
        upload = {
            "upload": {
                "name": "identity.pdf",
                "data": base64.b64encode(b"data").decode("ascii"),
            }
        }
        with self.assertRaisesRegex(APP.ApplicationError, "اسم الأب.*الشهرة"):
            APP.save_record(
                {
                    "mode": "create",
                    "record_code": "F0000001",
                    "main": {FLD_NAME: "علي", FLD_WORKS: "نعم"},
                    "related": {
                        CAT_DOCUMENTS: [
                            {
                                "values": {
                                    FLD_DOC_TYPE: "هوية",
                                    FLD_CUSTOM_FILE: upload,
                                }
                            }
                        ]
                    },
                }
            )
        self.assertEqual(APP.record_count(), 0)
        self.assertEqual(list(APP.attachments_directory().iterdir()), [])

    def test_13_delete_removes_every_related_row(self) -> None:
        self.configure()
        self.create_person(
            documents=[
                {"values": {FLD_DOC_TYPE: "واحد"}},
                {"values": {FLD_DOC_TYPE: "اثنان"}},
            ],
            children=[
                {"values": {FLD_CHILD_NAME: "أ"}},
                {"values": {FLD_CHILD_NAME: "ب"}},
            ],
        )
        result = APP.delete_record("A12F12B7")
        self.assertEqual(result["deleted_main_rows"], 1)
        self.assertEqual(result["deleted_related_rows"], 4)
        self.assertEqual(APP.record_count(), 0)
        with self.assertRaisesRegex(APP.ApplicationError, "لم يُعثر"):
            APP.load_record("A12F12B7")

    def test_14_schema_deletion_removes_values_and_orphan_files(self) -> None:
        self.configure()
        upload = {
            "upload": {
                "name": "a.pdf",
                "data": base64.b64encode(b"attachment").decode("ascii"),
            }
        }
        self.create_person(
            documents=[
                {
                    "values": {
                        FLD_DOC_TYPE: "هوية",
                        FLD_CUSTOM_FILE: upload,
                    }
                }
            ]
        )
        schema = APP.read_schema_file()
        schema["categories"] = [
            category
            for category in schema["categories"]
            if category["id"] != CAT_DOCUMENTS
        ]
        schema["conditions"] = [
            condition
            for condition in schema["conditions"]
            if condition["target_id"] != CAT_DOCUMENTS
        ]
        APP.save_schema(schema)
        loaded = APP.load_record("A12F12B7")
        self.assertNotIn(CAT_DOCUMENTS, loaded["related"])
        self.assertEqual(list(APP.attachments_directory().iterdir()), [])

    def test_15_concurrent_creates_do_not_corrupt_workbook(self) -> None:
        self.configure()
        barrier = threading.Barrier(10)

        def create(number: int) -> str:
            barrier.wait()
            result = APP.save_record(
                {
                    "mode": "create",
                    "record_code": valid_code(number),
                    "main": {FLD_NAME: f"شخص {number}"},
                    "related": {},
                }
            )
            return result["record_code"]

        with ThreadPoolExecutor(max_workers=10) as executor:
            codes = list(executor.map(create, range(10)))
        self.assertEqual(len(set(codes)), 10)
        self.assertEqual(APP.record_count(), 10)
        workbook = load_workbook(APP.WORKBOOK_PATH, read_only=True)
        workbook.close()
        self.assertEqual(list(self.root.glob(".database-*.xlsx")), [])

    def test_16_http_api_and_developer_only_builder(self) -> None:
        server = APP.DataEntryHTTPServer((APP.HOST, 0), APP.DataEntryRequestHandler)
        thread = threading.Thread(target=server.serve_forever, daemon=True)
        thread.start()
        base = f"http://{APP.HOST}:{server.server_address[1]}"

        def request(method: str, path: str, body: object | None = None):
            data = (
                json.dumps(body, ensure_ascii=False).encode("utf-8")
                if body is not None
                else None
            )
            req = Request(
                f"{base}{path}",
                data=data,
                headers={"Content-Type": "application/json"} if data else {},
                method=method,
            )
            with urlopen(req, timeout=5) as response:
                return response.status, json.load(response)

        try:
            status, health = request("GET", "/api/health")
            self.assertEqual(status, 200)
            self.assertTrue(health["ok"])
            self.assertFalse(health["browser_active"])
            status, schema = request("GET", "/api/schema")
            self.assertEqual(status, 200)
            self.assertTrue(schema["developer_mode"])

            APP.DEVELOPER_MODE = False
            with self.assertRaises(HTTPError) as blocked:
                request("PUT", "/api/schema", APP.read_schema_file())
            self.assertEqual(blocked.exception.code, 400)
            APP.DEVELOPER_MODE = True

            request("PUT", "/api/schema", APP.read_schema_file())
            configured = configured_schema(revision=1)
            status, saved = request("PUT", "/api/schema", configured)
            self.assertEqual(status, 200)
            self.assertEqual(saved["stats"]["field_count"], 18)
            status, backup = request("POST", "/api/backup")
            self.assertEqual(status, 200)
            self.assertTrue((APP.BACKUP_DIR / backup["filename"]).is_file())
            APP.DEVELOPER_MODE = False
            with self.assertRaises(HTTPError) as blocked_backup:
                request("POST", "/api/backup")
            self.assertEqual(blocked_backup.exception.code, 400)
            APP.DEVELOPER_MODE = True
            status, created = request(
                "POST",
                "/api/records",
                {
                    "mode": "create",
                    "record_code": "H0000001",
                    "main": {FLD_NAME: "HTTP"},
                    "related": {},
                },
            )
            self.assertEqual(status, 200)
            self.assertEqual(created["action"], "created")
            status, found = request("POST", "/api/search", {FLD_NAME: "http"})
            self.assertEqual(status, 200)
            self.assertEqual(found["matches"][0]["record_code"], "H0000001")
            self.assertEqual(
                request("GET", "/api/records/H0000001")[1]["main"][FLD_NAME],
                "HTTP",
            )
            self.assertEqual(
                request("DELETE", "/api/records/H0000001")[1]["deleted_main_rows"],
                1,
            )
        finally:
            server.shutdown()
            server.server_close()
            thread.join(timeout=5)

    def test_17_disconnect_does_not_terminate_server(self) -> None:
        server = APP.DataEntryHTTPServer((APP.HOST, 0), APP.DataEntryRequestHandler)
        thread = threading.Thread(target=server.serve_forever, daemon=True)
        thread.start()
        base = f"http://{APP.HOST}:{server.server_address[1]}"
        try:
            for path in ("/api/heartbeat", "/api/disconnect"):
                with urlopen(
                    Request(f"{base}{path}", data=b"", method="POST"),
                    timeout=5,
                ) as response:
                    self.assertEqual(response.status, 200)
            time.sleep(0.2)
            self.assertTrue(thread.is_alive())
            self.assertFalse(server.browser_is_active())
            with urlopen(f"{base}/api/health", timeout=5) as response:
                health = json.load(response)
            self.assertTrue(health["ok"])
            self.assertFalse(health["browser_active"])
        finally:
            server.shutdown()
            server.server_close()
            thread.join(timeout=5)

    def test_18_invalid_or_duplicate_ids_are_rejected(self) -> None:
        self.configure()
        self.create_person()
        with self.assertRaisesRegex(APP.ApplicationError, "مستخدم مسبق"):
            self.create_person()
        with self.assertRaisesRegex(APP.ApplicationError, "غير صالح"):
            self.create_person(code="123")
        loaded = APP.load_record("A12F12B7")
        child = loaded["related"][CAT_CHILDREN][0]
        with self.assertRaisesRegex(APP.ApplicationError, "لا يعود"):
            APP.save_record(
                {
                    "mode": "update",
                    "record_code": "A12F12B7",
                    "main": loaded["main"],
                    "related": {
                        CAT_CHILDREN: [
                            {
                                "_child_id": "f" * 32,
                                "values": child["values"],
                            }
                        ]
                    },
                }
            )

    def test_19_large_workbook_uses_fast_row_iteration(self) -> None:
        schema = self.configure()
        timestamp = APP.now_iso()
        records = [
            {
                "_record_id": f"{index + 1:032x}",
                "record_code": valid_code(index),
                "created_at": timestamp,
                "updated_at": timestamp,
                "values": {FLD_NAME: f"شخص {index}"},
                "related": {},
            }
            for index in range(1_000)
        ]
        APP.atomic_write_workbook(schema, records)
        started = time.perf_counter()
        loaded = APP.read_dataset(schema)
        elapsed = time.perf_counter() - started
        self.assertEqual(len(loaded), 1_000)
        self.assertLess(
            elapsed,
            5.0,
            f"Reading 1,000 records took {elapsed:.2f} seconds.",
        )

    def test_20_duplicate_excel_record_and_child_ids_are_rejected(self) -> None:
        schema = self.configure()
        self.create_person()

        workbook = load_workbook(APP.WORKBOOK_PATH)
        main = workbook[APP.MAIN_SHEET]
        duplicate = [cell.value for cell in main[APP.FIRST_DATA_ROW]]
        main.append(duplicate)
        workbook.save(APP.WORKBOOK_PATH)
        workbook.close()
        with self.assertRaisesRegex(APP.ApplicationError, "داخلي مكرر"):
            APP.read_dataset(schema)

        workbook = load_workbook(APP.WORKBOOK_PATH)
        main = workbook[APP.MAIN_SHEET]
        headers = APP.technical_headers(main)
        main.cell(
            main.max_row,
            headers.index("_record_id") + 1,
            "e" * 32,
        )
        workbook.save(APP.WORKBOOK_PATH)
        workbook.close()
        with self.assertRaisesRegex(APP.ApplicationError, "ظاهر.*مكرر"):
            APP.read_dataset(schema)

        workbook = load_workbook(APP.WORKBOOK_PATH)
        main = workbook[APP.MAIN_SHEET]
        main.delete_rows(main.max_row)
        child_sheet = workbook[APP.related_sheet_names(schema)[CAT_CHILDREN]]
        child_sheet.append(
            [cell.value for cell in child_sheet[APP.FIRST_DATA_ROW]]
        )
        workbook.save(APP.WORKBOOK_PATH)
        workbook.close()
        with self.assertRaisesRegex(APP.ApplicationError, "صف داخلي مكرر"):
            APP.read_dataset(schema)

    def test_21_schema_save_rolls_workbook_back_if_json_replace_fails(self) -> None:
        self.configure()
        self.create_person()
        original_schema = APP.SCHEMA_PATH.read_bytes()
        original_workbook = APP.WORKBOOK_PATH.read_bytes()
        schema = APP.read_schema_file()
        schema["app"]["title"] = "عنوان لن يكتمل"

        real_replace = APP.os.replace
        failed = False

        def fail_schema_once(source, destination):
            nonlocal failed
            if Path(destination) == APP.SCHEMA_PATH and not failed:
                failed = True
                raise PermissionError("simulated schema replacement failure")
            return real_replace(source, destination)

        with mock.patch.object(APP.os, "replace", side_effect=fail_schema_once):
            with self.assertRaisesRegex(APP.ApplicationError, "تعذّر تحديث الإعداد"):
                APP.save_schema(schema)

        self.assertTrue(failed)
        self.assertEqual(APP.SCHEMA_PATH.read_bytes(), original_schema)
        self.assertEqual(APP.WORKBOOK_PATH.read_bytes(), original_workbook)
        self.assertEqual(list(self.root.glob(".database-rollback-*")), [])

    def test_22_unsafe_attachments_download_instead_of_running_inline(self) -> None:
        self.configure()
        content = b"<script>document.body.textContent='unsafe'</script>"
        APP.attachments_directory().joinpath("unsafe.html").write_bytes(content)
        APP.attachments_directory().joinpath("safe.png").write_bytes(
            b"\x89PNG\r\n\x1a\n"
        )
        server = APP.DataEntryHTTPServer(
            (APP.HOST, 0),
            APP.DataEntryRequestHandler,
        )
        thread = threading.Thread(target=server.serve_forever, daemon=True)
        thread.start()
        base = f"http://{APP.HOST}:{server.server_address[1]}"
        try:
            with urlopen(f"{base}/api/attachments/unsafe.html") as response:
                self.assertEqual(response.read(), content)
                self.assertTrue(
                    response.headers["Content-Disposition"].startswith(
                        "attachment;"
                    )
                )
                self.assertEqual(
                    response.headers["Content-Type"],
                    "application/octet-stream",
                )
                self.assertEqual(
                    response.headers["Content-Security-Policy"],
                    "sandbox",
                )
            with urlopen(f"{base}/api/attachments/safe.png") as response:
                self.assertTrue(
                    response.headers["Content-Disposition"].startswith(
                        "inline;"
                    )
                )
                self.assertEqual(response.headers["Content-Type"], "image/png")
        finally:
            server.shutdown()
            server.server_close()
            thread.join(timeout=5)

    def test_23_backup_contains_schema_workbook_and_attachments(self) -> None:
        self.configure()
        self.create_person()
        attachment = APP.attachments_directory() / "مستند.txt"
        attachment.write_text("نسخة احتياطية", encoding="utf-8")
        result = APP.create_backup()
        backup = APP.backup_file_path(result["filename"])
        self.assertIsNotNone(backup)
        self.assertTrue(backup.is_file())
        with zipfile.ZipFile(backup) as archive:
            names = set(archive.namelist())
            self.assertIn("schema.json", names)
            self.assertIn("database.xlsx", names)
            self.assertIn("attachments/مستند.txt", names)
            self.assertEqual(
                archive.read("attachments/مستند.txt").decode("utf-8"),
                "نسخة احتياطية",
            )
        self.assertIsNone(APP.backup_file_path("../not-a-backup.zip"))


    def test_24_schema_v2_dependencies_markers_uniqueness_archive_and_validation(self) -> None:
        cat_main = "cat_100000000001"
        cat_education = "cat_100000000002"
        fld_employee = "fld_100000000001"
        fld_department = "fld_100000000002"
        fld_job = "fld_100000000003"
        fld_verified = "fld_100000000004"
        fld_skills = "fld_100000000005"
        fld_start = "fld_100000000006"
        fld_end = "fld_100000000007"
        fld_level = "fld_100000000008"
        marker_current = "mark_100000000001"
        dep_it = "opt_100000000001"
        dep_finance = "opt_100000000002"
        job_developer = "opt_100000000003"
        job_accountant = "opt_100000000004"
        skill_python = "opt_100000000005"
        skill_excel = "opt_100000000006"
        level_bachelor = "opt_100000000007"
        level_master = "opt_100000000008"

        schema = {
            "schema_version": 2,
            "revision": 0,
            "app": {
                "title": "نظام الموارد البشرية",
                "entity_singular": "موظف",
                "entity_plural": "الموظفون",
                "direction": "rtl",
                "primary_color": "#315F8A",
            },
            "categories": [
                {
                    "id": cat_main,
                    "label": "العمل",
                    "description": "",
                    "kind": "main",
                    "add_label": "",
                    "auto_start": False,
                    "anchor_field_id": None,
                    "row_markers": [],
                    "fields": [
                        {
                            **field(fld_employee, "رقم الموظف", "text", required=True, searchable=True),
                            "unique": True,
                            "validation": {"min_length": 3, "max_length": 10},
                        },
                        {
                            **field(fld_department, "القسم", "select"),
                            "options": [
                                {"id": dep_it, "label": "تقنية المعلومات", "active": True},
                                {"id": dep_finance, "label": "المالية", "active": True},
                            ],
                        },
                        {
                            **field(fld_job, "المسمى الوظيفي", "select"),
                            "options": [
                                {"id": job_developer, "label": "مبرمج", "active": True},
                                {"id": job_accountant, "label": "محاسب", "active": True},
                            ],
                            "option_filter": {
                                "source_field_id": fld_department,
                                "mappings": {
                                    dep_it: [job_developer],
                                    dep_finance: [job_accountant],
                                },
                                "unmatched": "none",
                            },
                        },
                        field(fld_verified, "تم التحقق", "checkbox", required=True),
                        {
                            **field(fld_skills, "المهارات", "checkbox_group"),
                            "options": [
                                {"id": skill_python, "label": "Python", "active": True},
                                {"id": skill_excel, "label": "Excel", "active": True},
                            ],
                        },
                        field(fld_start, "بداية العقد", "date_gregorian"),
                        {
                            **field(fld_end, "نهاية العقد", "date_gregorian"),
                            "validation": {
                                "compare_field_id": fld_start,
                                "compare_operator": "after",
                            },
                        },
                    ],
                },
                {
                    "id": cat_education,
                    "label": "المؤهلات",
                    "description": "",
                    "kind": "repeatable",
                    "add_label": "إضافة مؤهل",
                    "auto_start": False,
                    "anchor_field_id": None,
                    "row_markers": [
                        {
                            "id": marker_current,
                            "label": "الحالي",
                            "display_text": "# الحالي",
                            "color": "#0F766E",
                            "rule": "at_most_one",
                        }
                    ],
                    "fields": [
                        {
                            **field(fld_level, "المستوى", "select"),
                            "options": [
                                {"id": level_bachelor, "label": "بكالوريوس", "active": True},
                                {"id": level_master, "label": "ماجستير", "active": True},
                            ],
                        }
                    ],
                },
            ],
            "conditions": [],
        }
        saved = APP.save_schema(schema)
        self.assertEqual(saved["schema_version"], 2)

        APP.save_record(
            {
                "mode": "create",
                "record_code": "H0000001",
                "main": {
                    fld_employee: "EMP1",
                    fld_department: dep_it,
                    fld_job: job_developer,
                    fld_verified: True,
                    fld_skills: [skill_python, skill_excel],
                    fld_start: "2026-01-01",
                    fld_end: "2026-12-31",
                },
                "related": {
                    cat_education: [
                        {
                            "markers": {marker_current: True},
                            "values": {fld_level: level_master},
                        }
                    ]
                },
            }
        )
        loaded = APP.load_record("H0000001")
        self.assertEqual(loaded["main"][fld_department], "تقنية المعلومات")
        self.assertEqual(loaded["main"][fld_job], "مبرمج")
        self.assertTrue(loaded["main"][fld_verified])
        self.assertEqual(loaded["main"][fld_skills], ["Python", "Excel"])
        self.assertTrue(
            loaded["related"][cat_education][0]["markers"][marker_current]
        )

        with self.assertRaisesRegex(APP.ApplicationError, "غير متاحة"):
            APP.save_record(
                {
                    "mode": "create",
                    "record_code": "H0000002",
                    "main": {
                        fld_employee: "EMP2",
                        fld_department: dep_it,
                        fld_job: job_accountant,
                        fld_verified: True,
                    },
                    "related": {},
                }
            )

        with self.assertRaisesRegex(APP.ApplicationError, "مستخدمة"):
            APP.save_record(
                {
                    "mode": "create",
                    "record_code": "H0000003",
                    "main": {
                        fld_employee: "EMP1",
                        fld_department: dep_finance,
                        fld_job: job_accountant,
                        fld_verified: True,
                    },
                    "related": {},
                }
            )

        with self.assertRaisesRegex(APP.ApplicationError, "مرة واحدة"):
            APP.save_record(
                {
                    "mode": "create",
                    "record_code": "H0000004",
                    "main": {
                        fld_employee: "EMP4",
                        fld_department: dep_finance,
                        fld_job: job_accountant,
                        fld_verified: True,
                    },
                    "related": {
                        cat_education: [
                            {"markers": {marker_current: True}, "values": {fld_level: level_bachelor}},
                            {"markers": {marker_current: True}, "values": {fld_level: level_master}},
                        ]
                    },
                }
            )

        with self.assertRaisesRegex(APP.ApplicationError, "بعد"):
            APP.save_record(
                {
                    "mode": "create",
                    "record_code": "H0000005",
                    "main": {
                        fld_employee: "EMP5",
                        fld_department: dep_finance,
                        fld_job: job_accountant,
                        fld_verified: True,
                        fld_start: "2026-12-31",
                        fld_end: "2026-01-01",
                    },
                    "related": {},
                }
            )

        APP.archive_record("H0000001", True)
        self.assertEqual(
            APP.search_records({fld_employee: "EMP1"})["matches"], []
        )
        archived = APP.search_records(
            {fld_employee: "EMP1", "_include_archived": True}
        )["matches"]
        self.assertEqual(len(archived), 1)
        self.assertTrue(archived[0]["archived"])
        APP.archive_record("H0000001", False)
        self.assertFalse(APP.load_record("H0000001")["archived"])

    def test_25_builder_password_unlock_change_and_lock(self) -> None:
        APP.DEVELOPER_MODE = False
        APP._BUILDER_UNLOCKED = False
        self.assertFalse(APP.builder_access_response()["configured"])
        with self.assertRaisesRegex(APP.ApplicationError, "مقفل"):
            APP.require_builder_access()
        access = APP.unlock_builder("password-123", initialize=True)
        self.assertTrue(access["configured"])
        self.assertTrue(access["unlocked"])
        APP.require_builder_access()
        APP.lock_builder()
        self.assertFalse(APP.builder_access_response()["unlocked"])
        with self.assertRaisesRegex(APP.ApplicationError, "غير صحيحة"):
            APP.unlock_builder("wrong-password")
        APP.unlock_builder("password-123")
        access = APP.change_builder_password("", "new-password-456")
        self.assertTrue(access["unlocked"])
        APP.lock_builder()
        with self.assertRaisesRegex(APP.ApplicationError, "غير صحيحة"):
            APP.unlock_builder("password-123")
        self.assertTrue(APP.unlock_builder("new-password-456")["unlocked"])

    def test_26_optional_system_metadata_fields(self) -> None:
        schema = configured_schema()
        system_specs = (
            ("fld_100000000001", "معرّف السجل", "system_record_code"),
            ("fld_100000000002", "تاريخ الإنشاء", "system_created_at"),
            ("fld_100000000003", "تاريخ آخر تعديل", "system_updated_at"),
        )
        schema["categories"][0]["fields"].extend(
            field(field_id, label, field_type)
            for field_id, label, field_type in system_specs
        )
        saved = APP.save_schema(schema)
        self.assertEqual(saved["stats"]["field_count"], 21)

        workbook = load_workbook(APP.WORKBOOK_PATH)
        try:
            headers = APP.technical_headers(workbook[APP.MAIN_SHEET])
            self.assertFalse(
                {field_id for field_id, _, _ in system_specs} & set(headers)
            )
            self.assertTrue(
                {"record_code", "created_at", "updated_at"} <= set(headers)
            )
        finally:
            workbook.close()

        self.create_person()
        loaded = APP.load_record("A12F12B7")
        self.assertTrue(loaded["created_at"])
        self.assertTrue(loaded["updated_at"])
        self.assertFalse(
            {field_id for field_id, _, _ in system_specs} & set(loaded["main"])
        )

        duplicate = copy.deepcopy(saved)
        duplicate["categories"][0]["fields"].append(
            field("fld_100000000004", "معرّف مكرر", "system_record_code")
        )
        with self.assertRaisesRegex(APP.ApplicationError, "حقل واحد"):
            APP.validate_schema(duplicate)

        repeatable = copy.deepcopy(saved)
        next(
            category
            for category in repeatable["categories"]
            if category["id"] == CAT_CHILDREN
        )["fields"].append(
            field("fld_100000000005", "تاريخ تقني", "system_created_at")
        )
        with self.assertRaisesRegex(APP.ApplicationError, "السجل الرئيسي"):
            APP.validate_schema(repeatable)

    def test_27_related_person_mode_is_a_condition_source(self) -> None:
        schema = configured_schema()
        children = next(
            category
            for category in schema["categories"]
            if category["id"] == CAT_CHILDREN
        )
        children["related_person_enabled"] = True
        next(
            child_field
            for child_field in children["fields"]
            if child_field["id"] == FLD_CHILD_DATE
        )["required"] = True
        schema["conditions"].append(
            {
                "id": "cond_100000000001",
                "target_type": "field",
                "target_id": FLD_CHILD_DATE,
                "source_field_id": APP.related_person_mode_source_id(
                    CAT_CHILDREN
                ),
                "operator": "equals",
                "value": "manual",
            }
        )
        APP.save_schema(schema)

        APP.save_record(
            {
                "mode": "create",
                "record_code": "A12F12B7",
                "main": {FLD_NAME: "الشخص الأصلي"},
                "related": {},
            }
        )
        APP.save_record(
            {
                "mode": "create",
                "record_code": "B12F12B7",
                "main": {FLD_NAME: "سجل مرتبط"},
                "related": {
                    CAT_CHILDREN: [
                        {
                            "related_person_mode": "existing",
                            "linked_record_code": "A12F12B7",
                            "values": {},
                        }
                    ]
                },
            }
        )
        linked = APP.load_record("B12F12B7")["related"][CAT_CHILDREN][0]
        self.assertEqual(linked["linked_record_code"], "A12F12B7")

        with self.assertRaisesRegex(APP.ApplicationError, "مطلوب"):
            APP.save_record(
                {
                    "mode": "create",
                    "record_code": "C12F12B7",
                    "main": {FLD_NAME: "سجل يدوي"},
                    "related": {
                        CAT_CHILDREN: [
                            {
                                "related_person_mode": "manual",
                                "linked_record_code": "",
                                "values": {FLD_CHILD_NAME: "قيمة يدوية"},
                            }
                        ]
                    },
                }
            )

    def test_28_dataset_snapshot_reuses_workbook_and_detects_external_change(self) -> None:
        schema = self.configure()
        self.create_person()
        APP.invalidate_dataset_cache()

        with mock.patch.object(
            APP, "load_workbook", wraps=APP.load_workbook
        ) as workbook_loader:
            self.assertEqual(APP.load_record("A12F12B7")["main"][FLD_NAME], "عَلِي")
            self.assertEqual(APP.load_record("A12F12B7")["main"][FLD_NAME], "عَلِي")
            self.assertEqual(workbook_loader.call_count, 1)

        workbook = load_workbook(APP.WORKBOOK_PATH)
        try:
            worksheet = workbook[APP.MAIN_SHEET]
            headers = APP.technical_headers(worksheet)
            worksheet.cell(
                APP.FIRST_DATA_ROW,
                headers.index(FLD_NAME) + 1,
                "تعديل خارجي",
            )
            workbook.save(APP.WORKBOOK_PATH)
        finally:
            workbook.close()

        self.assertEqual(
            APP.load_record("A12F12B7")["main"][FLD_NAME],
            "تعديل خارجي",
        )

    def test_29_failed_save_never_mutates_published_snapshot(self) -> None:
        self.configure()
        self.create_person()
        loaded = APP.load_record("A12F12B7")
        changed_main = copy.deepcopy(loaded["main"])
        changed_main[FLD_NAME] = "قيمة لا يجب نشرها"

        with mock.patch.object(
            APP, "atomic_write_workbook", side_effect=PermissionError
        ):
            with self.assertRaisesRegex(APP.ApplicationError, "تعذّر الحفظ"):
                APP.save_record(
                    {
                        "mode": "update",
                        "record_code": "A12F12B7",
                        "main": changed_main,
                        "related": loaded["related"],
                    }
                )

        self.assertEqual(
            APP.load_record("A12F12B7")["main"][FLD_NAME],
            "عَلِي",
        )

    def test_30_public_dataset_reads_are_isolated_from_cache(self) -> None:
        schema = self.configure()
        self.create_person()
        records = APP.read_dataset(schema)
        records[0]["values"][FLD_NAME] = "تعديل في نسخة خارجية"
        records.clear()

        self.assertEqual(
            APP.load_record("A12F12B7")["main"][FLD_NAME],
            "عَلِي",
        )


if __name__ == "__main__":
    unittest.main(verbosity=2)
