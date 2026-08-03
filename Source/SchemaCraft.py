# -*- coding: utf-8 -*-
"""Portable, metadata-driven Excel data-entry application.

The distributed application starts with an empty schema.  Users define
categories, fields, search behaviour, visibility rules, and attachment naming
from Builder mode.  The schema uses stable IDs; Excel labels and sheet names are
presentation details and may change without breaking saved data.
"""

from __future__ import annotations

import base64
import binascii
import copy
import ctypes
import hashlib
import json
import logging
import mimetypes
import os
import re
import secrets
import shutil
import string
import subprocess
import sys
import tempfile
import threading
import time
import traceback
import unicodedata
import uuid
import webbrowser
import zipfile
from datetime import date, datetime
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import quote, unquote, urlparse
from urllib.error import URLError
from urllib.request import urlopen
from logging.handlers import RotatingFileHandler
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


HOST = "127.0.0.1"
APP_PORT_MIN = 51000
APP_PORT_SPAN = 10000
SCHEMA_VERSION = 2
SUPPORTED_SCHEMA_VERSIONS = {1, 2}
MAIN_SHEET = "السجلات"
META_SHEET = "_meta"
TECHNICAL_HEADER_ROW = 1
VISIBLE_HEADER_ROW = 2
FIRST_DATA_ROW = 3
BROWSER_ACTIVE_SECONDS = 90.0
STARTUP_BROWSER_WAIT_SECONDS = 5.0

MAIN_REQUIRED_INTERNAL_HEADERS = (
    "_record_id",
    "record_code",
    "created_at",
    "updated_at",
)
MAIN_INTERNAL_HEADERS = (
    *MAIN_REQUIRED_INTERNAL_HEADERS,
    "_archived",
    "_archived_at",
)
RELATED_INTERNAL_HEADERS = (
    "_child_id",
    "_record_id",
    "record_code",
    "minor_id",
    "created_at",
    "updated_at",
)
RELATED_LINK_HEADER = "_linked_record_code"
RELATED_PERSON_MODE_SOURCE_PREFIX = "related_person_mode:"

FIELD_TYPES = {
    "text",
    "textarea",
    "number",
    "select",
    "checkbox",
    "checkbox_group",
    "yes_no",
    "date_gregorian",
    "date_hijri",
    "date_persian",
    "file",
    "system_record_code",
    "system_created_at",
    "system_updated_at",
}
SYSTEM_FIELD_TYPES = {
    "system_record_code",
    "system_created_at",
    "system_updated_at",
}
CATEGORY_KINDS = {"main", "repeatable"}
FIELD_WIDTHS = {"normal", "wide", "full"}
SEARCH_MATCHES = {"contains", "exact"}
CONDITION_OPERATORS = {
    "equals",
    "not_equals",
    "contains",
    "not_contains",
    "greater_than",
    "greater_or_equal",
    "less_than",
    "less_or_equal",
    "before",
    "after",
    "on_or_before",
    "on_or_after",
    "empty",
    "not_empty",
}

CONDITION_OPERATORS_BY_FIELD_TYPE = {
    "text": {
        "equals",
        "not_equals",
        "contains",
        "not_contains",
        "empty",
        "not_empty",
    },
    "textarea": {
        "equals",
        "not_equals",
        "contains",
        "not_contains",
        "empty",
        "not_empty",
    },
    "number": {
        "equals",
        "not_equals",
        "greater_than",
        "greater_or_equal",
        "less_than",
        "less_or_equal",
        "empty",
        "not_empty",
    },
    "select": {
        "equals",
        "not_equals",
        "empty",
        "not_empty",
    },
    "yes_no": {
        "equals",
        "not_equals",
        "empty",
        "not_empty",
    },
    "checkbox": {
        "equals",
        "not_equals",
    },
    "checkbox_group": {
        "contains",
        "not_contains",
        "empty",
        "not_empty",
    },
    "date_gregorian": {
        "equals",
        "not_equals",
        "before",
        "after",
        "on_or_before",
        "on_or_after",
        "empty",
        "not_empty",
    },
    "date_hijri": {
        "equals",
        "not_equals",
        "before",
        "after",
        "on_or_before",
        "on_or_after",
        "empty",
        "not_empty",
    },
    "date_persian": {
        "equals",
        "not_equals",
        "before",
        "after",
        "on_or_before",
        "on_or_after",
        "empty",
        "not_empty",
    },
    "file": {
        "empty",
        "not_empty",
    },
}
DEFINITION_ID_PATTERN = re.compile(
    r"^(cat|fld|cond|grp|opt|mark)_[a-f0-9]{12}$"
)
PERSON_CODE_PATTERN = re.compile(r"^[A-Z][A-Z0-9]{7}$")
PERSON_CODE_FIRST_CHARACTERS = string.ascii_uppercase
PERSON_CODE_OTHER_CHARACTERS = string.ascii_uppercase + string.digits
INTERNAL_ID_PATTERN = re.compile(r"^[a-f0-9]{32}$")

MAX_ATTACHMENT_BYTES = 100 * 1024 * 1024
MAX_REQUEST_BYTES = 140 * 1024 * 1024
MAX_SEARCH_RESULTS = 100
EXCEL_DATE_FORMAT = "yyyy-mm-dd"
WINDOWS_ALREADY_EXISTS = 183

FIXED_BACKGROUND_COLOR = "#F4F7FB"
FIXED_SURFACE_COLOR = "#FFFFFF"
INLINE_ATTACHMENT_EXTENSIONS = {
    ".avif",
    ".bmp",
    ".gif",
    ".jpeg",
    ".jpg",
    ".pdf",
    ".png",
    ".webp",
}

ARABIC_DIACRITICS_PATTERN = re.compile(
    r"[\u0610-\u061A\u064B-\u065F\u0670\u06D6-\u06ED]"
)
ARABIC_SEARCH_TRANSLATION = {
    ord("أ"): "ا",
    ord("إ"): "ا",
    ord("آ"): "ا",
    ord("ٱ"): "ا",
    ord("ى"): "ي",
    ord("ـ"): "",
}
INVALID_WINDOWS_FILENAME_PATTERN = re.compile(r'[<>:"/\\|?*\x00-\x1f]')
INVALID_EXCEL_SHEET_PATTERN = re.compile(r"[\[\]:*?/\\]")
HEX_COLOR_PATTERN = re.compile(r"^#[0-9A-Fa-f]{6}$")

BASE_DIR = (
    Path(sys.executable).resolve().parent
    if getattr(sys, "frozen", False)
    else Path(__file__).resolve().parent
)
def application_port() -> int:
    """Use the same local port whenever this application folder starts."""
    folder_hash = hashlib.sha256(
        str(BASE_DIR).casefold().encode("utf-8")
    ).digest()

    return APP_PORT_MIN + (
        int.from_bytes(folder_hash[:2], "big") % APP_PORT_SPAN
    )
APP_DIR = BASE_DIR / "app"
DATA_DIR = BASE_DIR / "data"
LOG_DIR = DATA_DIR / "logs"
LOG_FILE = LOG_DIR / "application.log"
SCHEMA_PATH = DATA_DIR / "schema.json"
WORKBOOK_PATH = DATA_DIR / "database.xlsx"
WORKBOOK_LOCK = threading.RLock()


class DatasetSnapshot:
    """One coherent in-memory view of the Excel workbook and its indexes."""

    __slots__ = (
        "workbook_path",
        "workbook_signature",
        "schema_signature",
        "records",
        "records_by_code",
        "records_by_id",
        "record_positions",
        "unique_indexes",
        "search_main",
        "search_related",
    )

    def __init__(
        self,
        *,
        workbook_path: str,
        workbook_signature: tuple[int, int] | None,
        schema_signature: str,
        records: tuple[dict[str, Any], ...],
        records_by_code: dict[str, dict[str, Any]],
        records_by_id: dict[str, dict[str, Any]],
        record_positions: dict[str, int],
        unique_indexes: dict[str, dict[str, frozenset[str]]],
        search_main: dict[str, dict[str, Any]],
        search_related: dict[
            str, dict[str, tuple[dict[str, Any], ...]]
        ],
    ) -> None:
        self.workbook_path = workbook_path
        self.workbook_signature = workbook_signature
        self.schema_signature = schema_signature
        self.records = records
        self.records_by_code = records_by_code
        self.records_by_id = records_by_id
        self.record_positions = record_positions
        self.unique_indexes = unique_indexes
        self.search_main = search_main
        self.search_related = search_related


_DATASET_SNAPSHOT: DatasetSnapshot | None = None
DEVELOPER_ACCESS_PATH = BASE_DIR / "developer-access.key"
BACKUP_DIR = BASE_DIR / "backups"
BUILDER_AUTH_PATH = BASE_DIR / "builder-auth.json"
STARTUP_ERROR_LOG = BASE_DIR / "startup-error.log"
DEVELOPER_MODE = (
    "--builder" in sys.argv and DEVELOPER_ACCESS_PATH.is_file()
)
BUILDER_PASSWORD_ITERATIONS = 310_000
BUILDER_FAILED_AUTH_DELAY_SECONDS = 0.15
_BUILDER_SESSION_LOCK = threading.RLock()
_BUILDER_UNLOCKED = DEVELOPER_MODE
_WINDOWS_MUTEX_HANDLE: int | None = None
LOGGER = logging.getLogger("SchemaCraft")

class ApplicationError(Exception):
    """An error safe to display to the user."""
def configure_logging() -> None:
    LOG_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    LOGGER.setLevel(logging.INFO)
    LOGGER.propagate = False

    if LOGGER.handlers:
        return

    handler = RotatingFileHandler(
        LOG_FILE,
        maxBytes=5 * 1024 * 1024,
        backupCount=5,
        encoding="utf-8",
    )

    handler.setFormatter(
        logging.Formatter(
            "%(asctime)s | "
            "%(levelname)s | "
            "%(threadName)s | "
            "%(message)s"
        )
    )

    LOGGER.addHandler(handler)


def install_exception_logging() -> None:
    original_sys_hook = sys.excepthook
    original_thread_hook = threading.excepthook

    def application_exception_hook(
        exception_type,
        exception,
        traceback_object,
    ) -> None:
        if issubclass(exception_type, KeyboardInterrupt):
            original_sys_hook(
                exception_type,
                exception,
                traceback_object,
            )
            return

        LOGGER.critical(
            "Unhandled application exception",
            exc_info=(
                exception_type,
                exception,
                traceback_object,
            ),
        )

        original_sys_hook(
            exception_type,
            exception,
            traceback_object,
        )

    def thread_exception_hook(args) -> None:
        thread_name = (
            args.thread.name
            if args.thread is not None
            else "unknown"
        )

        LOGGER.critical(
            "Unhandled exception in thread %s",
            thread_name,
            exc_info=(
                args.exc_type,
                args.exc_value,
                args.exc_traceback,
            ),
        )

        original_thread_hook(args)

    sys.excepthook = application_exception_hook
    threading.excepthook = thread_exception_hook

def _read_builder_auth() -> dict[str, Any] | None:
    if not BUILDER_AUTH_PATH.is_file():
        return None
    try:
        payload = json.loads(BUILDER_AUTH_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise ApplicationError(f"تعذّر قراءة إعداد كلمة مرور المصمّم: {exc}") from exc
    if not isinstance(payload, dict):
        raise ApplicationError("إعداد كلمة مرور المصمّم غير صالح.")
    try:
        iterations = int(payload.get("iterations", 0))
        salt = base64.b64decode(str(payload.get("salt", "")), validate=True)
        digest = base64.b64decode(str(payload.get("hash", "")), validate=True)
    except (ValueError, binascii.Error) as exc:
        raise ApplicationError("إعداد كلمة مرور المصمّم تالف.") from exc
    if iterations < 100_000 or len(salt) < 16 or len(digest) != 32:
        raise ApplicationError("إعداد كلمة مرور المصمّم تالف.")
    return {"iterations": iterations, "salt": salt, "hash": digest}


def builder_password_configured() -> bool:
    return BUILDER_AUTH_PATH.is_file()


def _password_digest(password: str, salt: bytes, iterations: int) -> bytes:
    return hashlib.pbkdf2_hmac(
        "sha256",
        password.encode("utf-8"),
        salt,
        iterations,
        dklen=32,
    )


def verify_builder_password(password: Any) -> bool:
    text = str(password or "")
    auth = _read_builder_auth()
    if auth is None:
        return False
    candidate = _password_digest(text, auth["salt"], auth["iterations"])
    return secrets.compare_digest(candidate, auth["hash"])


def validate_new_builder_password(password: Any) -> str:
    text = str(password or "")
    if len(text) < 8:
        raise ApplicationError("يجب أن تتكون كلمة مرور المصمّم من 8 أحرف على الأقل.")
    if len(text) > 200:
        raise ApplicationError("كلمة مرور المصمّم طويلة جدًا.")
    return text


def set_builder_password(password: Any) -> None:
    text = validate_new_builder_password(password)
    salt = secrets.token_bytes(24)
    payload = {
        "version": 1,
        "algorithm": "pbkdf2-sha256",
        "iterations": BUILDER_PASSWORD_ITERATIONS,
        "salt": base64.b64encode(salt).decode("ascii"),
        "hash": base64.b64encode(
            _password_digest(text, salt, BUILDER_PASSWORD_ITERATIONS)
        ).decode("ascii"),
        "updated_at": now_iso(),
    }
    atomic_write_json(BUILDER_AUTH_PATH, payload)


def builder_is_unlocked() -> bool:
    if DEVELOPER_MODE:
        return True

    with _BUILDER_SESSION_LOCK:
        return _BUILDER_UNLOCKED


def unlock_builder(
    password: Any = None,
    *,
    initialize: bool = False,
) -> dict[str, Any]:
    global _BUILDER_UNLOCKED

    started_at = time.perf_counter()
    configured = builder_password_configured()

    try:
        if initialize:
            if configured:
                raise ApplicationError(
                    "تم إعداد كلمة مرور المصمّم مسبقًا."
                )
            set_builder_password(password)
        elif not configured:
            raise ApplicationError(
                "أنشئ كلمة مرور للمصمّم أولًا."
            )
        elif not verify_builder_password(password):
            time.sleep(BUILDER_FAILED_AUTH_DELAY_SECONDS)
            raise ApplicationError(
                "كلمة مرور المصمّم غير صحيحة."
            )

        with _BUILDER_SESSION_LOCK:
            _BUILDER_UNLOCKED = True

        return builder_access_response()
    finally:
        LOGGER.info(
            "Builder authentication completed in %.3f seconds",
            time.perf_counter() - started_at,
        )


def lock_builder() -> dict[str, Any]:
    global _BUILDER_UNLOCKED

    with _BUILDER_SESSION_LOCK:
        _BUILDER_UNLOCKED = False
    return builder_access_response()


def change_builder_password(current_password: Any, new_password: Any) -> dict[str, Any]:
    if builder_password_configured() and not builder_is_unlocked():
        if not verify_builder_password(current_password):
            time.sleep(BUILDER_FAILED_AUTH_DELAY_SECONDS)
            raise ApplicationError("كلمة المرور الحالية غير صحيحة.")
    set_builder_password(new_password)
    return unlock_builder(new_password)


def require_builder_access() -> None:
    if not builder_is_unlocked():
        raise ApplicationError("المصمّم مقفل. أدخل كلمة المرور أولًا.")


def builder_access_response() -> dict[str, Any]:
    return {
        "configured": builder_password_configured(),
        "unlocked": builder_is_unlocked(),
        "session_minutes": 0,
        "session_timeout_enabled": False,
        "developer_key_available": DEVELOPER_ACCESS_PATH.is_file(),
    }


def default_schema() -> dict[str, Any]:
    return {
        "schema_version": SCHEMA_VERSION,
        "revision": 0,
        "app": {
            "title": "نظام إدخال البيانات",
            "entity_singular": "سجل",
            "entity_plural": "السجلات",
            "direction": "rtl",
            "primary_color": "#1F5F95",
            "background_color": FIXED_BACKGROUND_COLOR,
            "surface_color": FIXED_SURFACE_COLOR,
        },
        "categories": [],
        "conditions": [],
    }


def new_definition_id(prefix: str) -> str:
    return f"{prefix}_{secrets.token_hex(6)}"


def new_internal_id() -> str:
    return uuid.uuid4().hex


def now_iso() -> str:
    return datetime.now().astimezone().isoformat(timespec="seconds")


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_search_text(value: Any) -> str:
    text = unicodedata.normalize("NFKC", str(json_value(value)))
    text = ARABIC_DIACRITICS_PATTERN.sub("", text)
    text = text.translate(ARABIC_SEARCH_TRANSLATION)
    return " ".join(text.casefold().split())


def json_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, datetime):
        if value.time().isoformat() == "00:00:00":
            return value.date().isoformat()
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


def validate_person_code(value: Any) -> str:
    code = clean_text(value)
    if not PERSON_CODE_PATTERN.fullmatch(code):
        raise ApplicationError("معرّف السجل غير صالح.")
    return code


def generate_person_code() -> str:
    return secrets.choice(PERSON_CODE_FIRST_CHARACTERS) + "".join(
        secrets.choice(PERSON_CODE_OTHER_CHARACTERS) for _ in range(7)
    )


def validate_internal_id(value: Any, label: str) -> str:
    identifier = clean_text(value)
    if not INTERNAL_ID_PATTERN.fullmatch(identifier):
        raise ApplicationError(f"{label} الداخلي غير صالح.")
    return identifier


def is_persian_leap_year(year: int) -> bool:
    return year % 33 in {1, 5, 9, 13, 17, 22, 26, 30}


def validate_persian_date(value: str, label: str) -> str:
    if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", value):
        raise ApplicationError(f'قيمة الحقل "{label}" ليست تاريخًا شمسيًا صالحًا.')
    year, month, day = (int(part) for part in value.split("-"))
    if not 1 <= month <= 12:
        raise ApplicationError(f'قيمة الحقل "{label}" ليست تاريخًا شمسيًا صالحًا.')
    maximum_day = (
        31
        if month <= 6
        else 30
        if month <= 11
        else 30
        if is_persian_leap_year(year)
        else 29
    )
    if not 1 <= day <= maximum_day:
        raise ApplicationError(f'قيمة الحقل "{label}" ليست تاريخًا شمسيًا صالحًا.')
    return value


def is_hijri_leap_year(year: int) -> bool:
    """Return the arithmetic Hijri leap-year result for validation.

    Actual observed month starts can differ by one day.  The application stores
    the user's chosen Hijri date as entered; this validation only prevents
    impossible selector combinations.
    """

    return year % 30 in {2, 5, 7, 10, 13, 16, 18, 21, 24, 26, 29}


def validate_hijri_date(value: str, label: str) -> str:
    if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", value):
        raise ApplicationError(f'قيمة الحقل "{label}" ليست تاريخًا هجريًا صالحًا.')
    year, month, day = (int(part) for part in value.split("-"))
    if not 1 <= month <= 12:
        raise ApplicationError(f'قيمة الحقل "{label}" ليست تاريخًا هجريًا صالحًا.')
    maximum_day = (
        30
        if month % 2 == 1
        else 30
        if month == 12 and is_hijri_leap_year(year)
        else 29
    )
    if not 1 <= day <= maximum_day:
        raise ApplicationError(f'قيمة الحقل "{label}" ليست تاريخًا هجريًا صالحًا.')
    return value


def normalize_field_value(value: Any, field: dict[str, Any]) -> Any:
    field_type = field["type"]

    if field_type == "file":
        return value
    if field_type == "checkbox":
        if isinstance(value, bool):
            return value
        text = clean_text(value).casefold()
        if text in {"", "false", "0", "لا", "no", "off"}:
            return False
        if text in {"true", "1", "نعم", "yes", "on"}:
            return True
        raise ApplicationError(f'قيمة الحقل "{field["label"]}" ليست اختيارًا صحيحًا.')
    if field_type == "checkbox_group":
        result: list[str] = []
        for item in parse_checkbox_group_value(value):
            label = normalize_choice_value(item, field)
            if label not in result:
                result.append(label)
        return result

    if value is None:
        return ""
    if isinstance(value, str):
        value = value.strip()
    if value == "":
        return ""

    if field_type == "date_gregorian":
        if isinstance(value, datetime):
            normalized: Any = value.date()
        elif isinstance(value, date):
            normalized = value
        else:
            try:
                normalized = date.fromisoformat(str(value))
            except ValueError as exc:
                raise ApplicationError(
                    f'قيمة الحقل "{field["label"]}" ليست تاريخًا ميلاديًا صالحًا.'
                ) from exc
        _validate_simple_constraints(normalized, field)
        return normalized

    if field_type == "date_persian":
        normalized = validate_persian_date(str(value), field["label"])
        _validate_simple_constraints(normalized, field)
        return normalized

    if field_type == "date_hijri":
        normalized = validate_hijri_date(str(value), field["label"])
        _validate_simple_constraints(normalized, field)
        return normalized

    if field_type == "number":
        if isinstance(value, bool):
            raise ApplicationError(f'قيمة الحقل "{field["label"]}" ليست رقمًا صالحًا.')
        try:
            number = float(value)
        except (TypeError, ValueError) as exc:
            raise ApplicationError(
                f'قيمة الحقل "{field["label"]}" ليست رقمًا صالحًا.'
            ) from exc
        normalized = int(number) if number.is_integer() else number
        _validate_simple_constraints(normalized, field)
        return normalized

    if field_type == "yes_no":
        normalized = normalize_choice_value(value, field)
        _validate_simple_constraints(normalized, field)
        return normalized

    if field_type == "select":
        normalized = normalize_choice_value(value, field)
        _validate_simple_constraints(normalized, field)
        return normalized

    normalized = str(value)
    _validate_simple_constraints(normalized, field)
    return normalized


def _validate_simple_constraints(value: Any, field: dict[str, Any]) -> None:
    validation = field.get("validation", {})
    if not validation or value in {"", None}:
        return
    label = field["label"]
    field_type = field["type"]
    if field_type in {"text", "textarea"}:
        length = len(str(value))
        minimum = validation.get("min_length")
        maximum = validation.get("max_length")
        if minimum is not None and length < minimum:
            raise ApplicationError(f'الحقل "{label}" يجب ألا يقل عن {minimum} أحرف.')
        if maximum is not None and length > maximum:
            raise ApplicationError(f'الحقل "{label}" يجب ألا يزيد على {maximum} أحرف.')
        pattern = validation.get("pattern")
        if pattern and re.fullmatch(pattern, str(value)) is None:
            raise ApplicationError(f'قيمة الحقل "{label}" لا تطابق النمط المطلوب.')
    elif field_type == "number":
        if validation.get("integer_only") and not float(value).is_integer():
            raise ApplicationError(f'الحقل "{label}" يقبل عددًا صحيحًا فقط.')
        minimum = validation.get("min")
        maximum = validation.get("max")
        if minimum is not None and float(value) < minimum:
            raise ApplicationError(f'قيمة الحقل "{label}" يجب ألا تقل عن {minimum}.')
        if maximum is not None and float(value) > maximum:
            raise ApplicationError(f'قيمة الحقل "{label}" يجب ألا تزيد على {maximum}.')
    elif field_type.startswith("date_"):
        current = value.isoformat() if isinstance(value, date) else str(value)
        minimum = validation.get("min_date")
        maximum = validation.get("max_date")
        if minimum and current < minimum:
            raise ApplicationError(f'تاريخ الحقل "{label}" أقدم من الحد المسموح.')
        if maximum and current > maximum:
            raise ApplicationError(f'تاريخ الحقل "{label}" أحدث من الحد المسموح.')


def _comparison_value(value: Any) -> str:
    if isinstance(value, date):
        return value.isoformat()
    return clean_text(value)


def validate_cross_field_constraints(
    fields: Iterable[dict[str, Any]], values: dict[str, Any]
) -> None:
    lookup = {field["id"]: field for field in fields}
    for field in lookup.values():
        validation = field.get("validation", {})
        other_id = validation.get("compare_field_id")
        operator = validation.get("compare_operator")
        if not other_id or not operator or other_id not in lookup:
            continue
        left = _comparison_value(values.get(field["id"], ""))
        right = _comparison_value(values.get(other_id, ""))
        if not left or not right:
            continue
        valid = {
            "before": left < right,
            "after": left > right,
            "on_or_before": left <= right,
            "on_or_after": left >= right,
        }.get(operator, True)
        if not valid:
            other = lookup[other_id]
            phrases = {
                "before": "قبل",
                "after": "بعد",
                "on_or_before": "في أو قبل",
                "on_or_after": "في أو بعد",
            }
            raise ApplicationError(
                f'يجب أن يكون الحقل "{field["label"]}" {phrases[operator]} '
                f'الحقل "{other["label"]}".'
            )


def excel_value(value: Any, field: dict[str, Any]) -> Any:
    normalized = normalize_field_value(value, field)
    if field["type"] == "checkbox_group":
        return " | ".join(normalized)
    if field["type"] == "checkbox":
        return "نعم" if normalized else "لا"
    return normalized



def _require_definition_id(value: Any, prefix: str, label: str) -> str:
    identifier = clean_text(value) or new_definition_id(prefix)
    if (
        not DEFINITION_ID_PATTERN.fullmatch(identifier)
        or not identifier.startswith(f"{prefix}_")
    ):
        raise ApplicationError(f"{label} الداخلي غير صالح.")
    return identifier


def _stable_option_id(field_id: str, label: str) -> str:
    digest = hashlib.sha256(f"{field_id}\0{label}".encode("utf-8")).hexdigest()[:12]
    return f"opt_{digest}"


def _normalize_options(values: Any, field_id: str) -> list[dict[str, Any]]:
    if not isinstance(values, list):
        return []
    result: list[dict[str, Any]] = []
    ids: set[str] = set()
    labels: set[str] = set()
    for value in values:
        if isinstance(value, dict):
            label = clean_text(value.get("label"))
            option_id = clean_text(value.get("id")) or _stable_option_id(field_id, label)
            active = bool(value.get("active", True))
        else:
            label = clean_text(value)
            option_id = _stable_option_id(field_id, label)
            active = True
        if not label:
            continue
        normalized_label = normalize_search_text(label)
        if normalized_label in labels:
            continue
        option_id = _require_definition_id(option_id, "opt", "معرّف الخيار")
        if option_id in ids:
            raise ApplicationError("يوجد معرّف خيار مكرر.")
        ids.add(option_id)
        labels.add(normalized_label)
        result.append({"id": option_id, "label": label, "active": active})
    return result


def option_labels(field: dict[str, Any], *, active_only: bool = False) -> list[str]:
    return [
        option["label"]
        for option in field.get("options", [])
        if not active_only or option.get("active", True)
    ]


def option_by_id(field: dict[str, Any], option_id: Any) -> dict[str, Any] | None:
    text = clean_text(option_id)
    return next(
        (option for option in field.get("options", []) if option["id"] == text),
        None,
    )


def option_by_value(field: dict[str, Any], value: Any) -> dict[str, Any] | None:
    text = clean_text(value)
    return next(
        (
            option
            for option in field.get("options", [])
            if option["id"] == text or option["label"] == text
        ),
        None,
    )


def option_token(field: dict[str, Any], value: Any) -> str:
    if field.get("type") == "checkbox":
        if isinstance(value, bool):
            return "true" if value else "false"
        return "true" if clean_text(value).casefold() in {"true", "1", "نعم", "yes"} else "false"
    option = option_by_value(field, value)
    return option["id"] if option else clean_text(value)


def normalize_choice_value(value: Any, field: dict[str, Any]) -> str:
    option = option_by_value(field, value)
    if option is None or not option.get("active", True):
        raise ApplicationError(
            f'قيمة الحقل "{field["label"]}" ليست ضمن الخيارات المسموحة.'
        )
    return option["label"]


def parse_checkbox_group_value(value: Any) -> list[Any]:
    if value is None or value == "":
        return []
    if isinstance(value, list):
        return value
    if isinstance(value, tuple):
        return list(value)
    text = clean_text(value)
    if not text:
        return []
    if text.startswith("["):
        try:
            parsed = json.loads(text)
            if isinstance(parsed, list):
                return parsed
        except json.JSONDecodeError:
            pass
    return [part.strip() for part in text.split(" | ") if part.strip()]




def _optional_int(value: Any, label: str, *, minimum: int = 0) -> int | None:
    if value in {None, ""}:
        return None
    try:
        number = int(value)
    except (TypeError, ValueError) as exc:
        raise ApplicationError(f"{label} يجب أن يكون عددًا صحيحًا.") from exc
    if number < minimum:
        raise ApplicationError(f"{label} يجب ألا يقل عن {minimum}.")
    return number


def _optional_float(value: Any, label: str) -> float | None:
    if value in {None, ""}:
        return None
    try:
        return float(value)
    except (TypeError, ValueError) as exc:
        raise ApplicationError(f"{label} يجب أن يكون رقمًا.") from exc


def _normalize_validation(raw: Any, field_type: str, field_label: str) -> dict[str, Any]:
    raw = raw if isinstance(raw, dict) else {}
    result: dict[str, Any] = {}
    if field_type in {"text", "textarea"}:
        minimum = _optional_int(raw.get("min_length"), f'الحد الأدنى للحقل "{field_label}"')
        maximum = _optional_int(raw.get("max_length"), f'الحد الأقصى للحقل "{field_label}"')
        if minimum is not None and maximum is not None and minimum > maximum:
            raise ApplicationError(f'حدود الطول للحقل "{field_label}" غير صحيحة.')
        pattern = clean_text(raw.get("pattern"))
        if pattern:
            try:
                re.compile(pattern)
            except re.error as exc:
                raise ApplicationError(f'نمط التحقق للحقل "{field_label}" غير صالح: {exc}') from exc
        result.update({"min_length": minimum, "max_length": maximum, "pattern": pattern})
    elif field_type == "number":
        minimum = _optional_float(raw.get("min"), f'الحد الأدنى للحقل "{field_label}"')
        maximum = _optional_float(raw.get("max"), f'الحد الأقصى للحقل "{field_label}"')
        if minimum is not None and maximum is not None and minimum > maximum:
            raise ApplicationError(f'حدود الرقم للحقل "{field_label}" غير صحيحة.')
        result.update({"min": minimum, "max": maximum, "integer_only": bool(raw.get("integer_only"))})
    elif field_type.startswith("date_"):
        result.update(
            {
                "min_date": clean_text(raw.get("min_date")),
                "max_date": clean_text(raw.get("max_date")),
                "compare_field_id": clean_text(raw.get("compare_field_id")) or None,
                "compare_operator": raw.get("compare_operator")
                if raw.get("compare_operator") in {"before", "after", "on_or_before", "on_or_after"}
                else None,
            }
        )
    return result


def _normalize_row_markers(values: Any, category_label: str) -> list[dict[str, Any]]:
    if not isinstance(values, list):
        return []
    result: list[dict[str, Any]] = []
    ids: set[str] = set()
    labels: set[str] = set()
    for raw in values:
        if not isinstance(raw, dict):
            continue
        marker_id = _require_definition_id(raw.get("id"), "mark", "معرّف وسم البطاقة")
        label = clean_text(raw.get("label"))
        if not label:
            raise ApplicationError(f'اسم وسم البطاقة مطلوب في الفئة "{category_label}".')
        normalized = normalize_search_text(label)
        if marker_id in ids or normalized in labels:
            raise ApplicationError(f'يوجد وسم بطاقة مكرر في الفئة "{category_label}".')
        ids.add(marker_id)
        labels.add(normalized)
        rule = raw.get("rule", "independent")
        if rule not in {"independent", "at_most_one", "exactly_one_when_rows", "exactly_one_always"}:
            rule = "independent"
        color = clean_text(raw.get("color"))
        if not HEX_COLOR_PATTERN.fullmatch(color):
            color = "#0F766E"
        result.append(
            {
                "id": marker_id,
                "label": label,
                "display_text": clean_text(raw.get("display_text")) or f"# {label}",
                "color": color,
                "rule": rule,
            }
        )
    return result


def condition_operators_for_field(
    field_type: str,
) -> set[str]:
    return CONDITION_OPERATORS_BY_FIELD_TYPE.get(
        field_type,
        {"equals", "not_equals", "empty", "not_empty"},
    )


def stable_condition_group_id(
    target_type: str,
    target_id: str,
) -> str:
    digest = hashlib.sha256(
        f"{target_type}\0{target_id}".encode("utf-8")
    ).hexdigest()[:12]

    return f"grp_{digest}"


def related_person_mode_source_id(category_id: str) -> str:
    return f"{RELATED_PERSON_MODE_SOURCE_PREFIX}{category_id}"


def related_person_mode_source_field(
    category: dict[str, Any],
) -> dict[str, Any] | None:
    if (
        category.get("kind") != "repeatable"
        or not category.get("related_person_enabled")
    ):
        return None
    return {
        "id": related_person_mode_source_id(category["id"]),
        "label": "هل لديه سجل؟",
        "type": "yes_no",
        "options": [
            {"id": "existing", "label": "لديه سجل", "active": True},
            {"id": "manual", "label": "ليس لديه سجل", "active": True},
        ],
    }


def normalize_condition_value(
    source_field: dict[str, Any],
    operator: str,
    raw_value: Any,
) -> str:
    if operator in {"empty", "not_empty"}:
        return ""

    field_type = source_field["type"]
    value = clean_text(raw_value)

    if not value:
        raise ApplicationError(
            f'أدخل قيمة شرط الحقل "{source_field["label"]}".'
        )

    if field_type == "checkbox":
        normalized = value.casefold()

        if normalized in {
            "true",
            "1",
            "yes",
            "نعم",
            "checked",
            "محدد",
        }:
            return "true"

        if normalized in {
            "false",
            "0",
            "no",
            "لا",
            "unchecked",
            "غير محدد",
        }:
            return "false"

        raise ApplicationError(
            f'قيمة شرط مربع الاختيار '
            f'"{source_field["label"]}" غير صالحة.'
        )

    if field_type in {
        "select",
        "yes_no",
        "checkbox_group",
    }:
        option = option_by_value(
            source_field,
            value,
        )

        if option is None:
            raise ApplicationError(
                f'قيمة شرط الحقل '
                f'"{source_field["label"]}" '
                "ليست ضمن خياراته."
            )

        # Store the stable option ID, not its editable label.
        return option["id"]

    if field_type == "number":
        try:
            number = float(value)
        except ValueError as exc:
            raise ApplicationError(
                f'قيمة شرط الحقل '
                f'"{source_field["label"]}" ليست رقمًا.'
            ) from exc

        return (
            str(int(number))
            if number.is_integer()
            else format(number, ".15g")
        )

    if field_type == "date_gregorian":
        try:
            return date.fromisoformat(value).isoformat()
        except ValueError as exc:
            raise ApplicationError(
                f'قيمة شرط التاريخ '
                f'"{source_field["label"]}" غير صالحة.'
            ) from exc

    if field_type == "date_hijri":
        return validate_hijri_date(
            value,
            source_field["label"],
        )

    if field_type == "date_persian":
        return validate_persian_date(
            value,
            source_field["label"],
        )

    return value

def validate_schema(payload: Any) -> dict[str, Any]:
    if not isinstance(payload, dict):
        raise ApplicationError("صيغة إعداد التطبيق غير صحيحة.")

    app_payload = payload.get("app")
    app_payload = app_payload if isinstance(app_payload, dict) else {}
    app = {
        "title": clean_text(app_payload.get("title")) or "نظام إدخال البيانات",
        "entity_singular": clean_text(app_payload.get("entity_singular")) or "سجل",
        "entity_plural": clean_text(app_payload.get("entity_plural")) or "السجلات",
        "direction": "rtl",
        "primary_color": (
            app_payload.get("primary_color")
            if HEX_COLOR_PATTERN.fullmatch(
                clean_text(app_payload.get("primary_color"))
            )
            else "#1F5F95"
        ),
        # These keys remain in Schema v1 for compatibility.  Their values are
        # fixed so a custom theme cannot hide controls or their text.
        "background_color": FIXED_BACKGROUND_COLOR,
        "surface_color": FIXED_SURFACE_COLOR,
    }

    raw_categories = payload.get("categories", [])
    if not isinstance(raw_categories, list):
        raise ApplicationError("قائمة الفئات غير صحيحة.")

    categories: list[dict[str, Any]] = []
    category_ids: set[str] = set()
    category_labels: set[str] = set()
    field_ids: set[str] = set()
    field_lookup: dict[str, dict[str, Any]] = {}
    field_category: dict[str, str] = {}

    for raw_category in raw_categories:
        if not isinstance(raw_category, dict):
            raise ApplicationError("أحد تعريفات الفئات غير صحيح.")
        category_id = _require_definition_id(
            raw_category.get("id"), "cat", "معرّف الفئة"
        )
        if category_id in category_ids:
            raise ApplicationError("يوجد معرّف فئة مكرر.")
        category_ids.add(category_id)

        label = clean_text(raw_category.get("label"))
        if not label:
            raise ApplicationError("اسم الفئة مطلوب.")
        normalized_label = normalize_search_text(label)
        if normalized_label in category_labels:
            raise ApplicationError(f'اسم الفئة "{label}" مكرر.')
        category_labels.add(normalized_label)

        kind = raw_category.get("kind", "main")
        if kind not in CATEGORY_KINDS:
            raise ApplicationError(f'نوع الفئة "{label}" غير صالح.')

        related_person_enabled = kind == "repeatable" and bool(
            raw_category.get("related_person_enabled")
        )

        raw_fields = raw_category.get("fields", [])
        if not isinstance(raw_fields, list):
            raise ApplicationError(f'حقول الفئة "{label}" غير صحيحة.')

        fields: list[dict[str, Any]] = []
        local_labels: set[str] = set()
        for raw_field in raw_fields:
            if not isinstance(raw_field, dict):
                raise ApplicationError(f'أحد حقول الفئة "{label}" غير صحيح.')
            field_id = _require_definition_id(
                raw_field.get("id"), "fld", "معرّف الحقل"
            )
            if field_id in field_ids:
                raise ApplicationError("يوجد معرّف حقل مكرر.")
            field_ids.add(field_id)

            field_label = clean_text(raw_field.get("label"))
            if not field_label:
                raise ApplicationError(f'يوجد حقل بلا اسم في الفئة "{label}".')
            normalized_field_label = normalize_search_text(field_label)
            if normalized_field_label in local_labels:
                raise ApplicationError(
                    f'اسم الحقل "{field_label}" مكرر داخل الفئة "{label}".'
                )
            local_labels.add(normalized_field_label)

            field_type = raw_field.get("type", "text")
            if field_type not in FIELD_TYPES:
                raise ApplicationError(f'نوع الحقل "{field_label}" غير صالح.')
            if field_type in SYSTEM_FIELD_TYPES and kind != "main":
                raise ApplicationError(
                    f'الحقل "{field_label}" يعرض بيانات السجل الرئيسي فقط.'
                )
            options = _normalize_options(raw_field.get("options", []), field_id)
            if field_type == "yes_no":
                existing = {option["label"]: option for option in options}
                options = [
                    existing.get(label)
                    or {
                        "id": _stable_option_id(field_id, label),
                        "label": label,
                        "active": True,
                    }
                    for label in ("نعم", "لا")
                ]
            if field_type in {"select", "checkbox_group"} and not any(
                option.get("active", True) for option in options
            ):
                raise ApplicationError(
                    f'أضف خيارًا واحدًا على الأقل للحقل "{field_label}".'
                )

            system_field = field_type in SYSTEM_FIELD_TYPES
            searchable = (
                bool(raw_field.get("searchable"))
                and field_type != "file"
                and not system_field
            )
            show_in_results = (
                bool(raw_field.get("show_in_results"))
                and field_type != "file"
                and not system_field
            )
            result_title = (
                bool(raw_field.get("result_title"))
                and show_in_results
                and kind == "main"
            )
            default_match = (
                "contains"
                if field_type in {"text", "textarea"}
                else "exact"
            )
            search_match = raw_field.get("search_match", default_match)
            if search_match not in SEARCH_MATCHES:
                search_match = default_match

            field = {
                "id": field_id,
                "label": field_label,
                "type": field_type,
                "required": bool(raw_field.get("required")) and not system_field,
                "placeholder": (
                    "" if system_field else clean_text(raw_field.get("placeholder"))
                ),
                "width": (
                    raw_field.get("width")
                    if raw_field.get("width") in FIELD_WIDTHS
                    else "normal"
                ),
                "options": options,
                "searchable": searchable,
                "search_match": search_match,
                "show_in_results": show_in_results,
                "result_title": result_title,
                "unique": (
                    bool(raw_field.get("unique"))
                    and field_type not in {"file", "checkbox"}
                    and not system_field
                ),
                "validation": (
                    {}
                    if system_field
                    else _normalize_validation(
                        raw_field.get("validation"), field_type, field_label
                    )
                ),
                "option_filter": None,
                "_option_filter_raw": raw_field.get("option_filter"),
                "related_person_source_field_id": (
                    clean_text(raw_field.get("related_person_source_field_id"))
                    or None
                    if related_person_enabled
                    else None
                ),
            }

            if field_type == "file":
                field["image_display"] = (
                    "profile"
                    if kind == "main"
                    and raw_field.get("image_display") == "profile"
                    else None
                )
                raw_naming = raw_field.get("file_naming")
                raw_naming = raw_naming if isinstance(raw_naming, dict) else {}
                mode = raw_naming.get("mode", "original")
                if mode not in {"original", "template"}:
                    mode = "original"
                raw_parts = raw_naming.get("parts", [])
                parts = []
                if isinstance(raw_parts, list):
                    for raw_part in raw_parts:
                        if not isinstance(raw_part, dict):
                            continue
                        source_field_id = clean_text(raw_part.get("field_id"))
                        if source_field_id:
                            parts.append(
                                {
                                    "field_id": source_field_id,
                                    "prefix": str(
                                        raw_part.get("prefix", "")
                                    )[:30],
                                    "suffix": str(
                                        raw_part.get("suffix", "")
                                    )[:30],
                                }
                            )
                field["file_naming"] = {"mode": mode, "parts": parts}

            fields.append(field)
            field_lookup[field_id] = field
            field_category[field_id] = category_id

        anchor_field_id = clean_text(raw_category.get("anchor_field_id")) or None
        parent_category_id = (
            clean_text(raw_category.get("parent_category_id")) or None
        )
        categories.append(
            {
                "id": category_id,
                "label": label,
                "description": clean_text(raw_category.get("description")),
                "kind": kind,
                "add_label": (
                    clean_text(raw_category.get("add_label"))
                    or f"إضافة {label}"
                ),
                "auto_start": bool(raw_category.get("auto_start")),
                "anchor_field_id": anchor_field_id,
                "parent_category_id": parent_category_id,
                "related_person_enabled": related_person_enabled,
                "row_markers": (
                    _normalize_row_markers(raw_category.get("row_markers", []), label)
                    if kind == "repeatable"
                    else []
                ),
                "fields": fields,
            }
        )

    category_lookup = {category["id"]: category for category in categories}

    for category in categories:
        parent_id = category["parent_category_id"]
        if parent_id and parent_id not in category_lookup:
            raise ApplicationError(
                f'الفئة الأم للفئة "{category["label"]}" محذوفة.'
            )
        if parent_id == category["id"]:
            raise ApplicationError("لا يمكن أن تكون الفئة أمًّا لنفسها.")

    for category in categories:
        visited = {category["id"]}
        current_id = category["parent_category_id"]
        while current_id:
            if current_id in visited:
                raise ApplicationError("يوجد تسلسل دائري بين الفئات الأم والفرعية.")
            visited.add(current_id)
            current_id = category_lookup[current_id]["parent_category_id"]

    for category in categories:
        if not category.get("related_person_enabled"):
            continue
        for field in category["fields"]:
            source_id = field.get("related_person_source_field_id")
            if not source_id:
                continue
            source = field_lookup.get(source_id)
            source_category = category_lookup.get(field_category.get(source_id, ""))
            if not source or not source_category or source_category["kind"] != "main":
                raise ApplicationError(
                    f'حقل النسخ المرتبط بالحقل "{field["label"]}" غير متاح.'
                )
            if source["type"] == "file" or source["type"] != field["type"]:
                raise ApplicationError(
                    f'نوع حقل المصدر لا يطابق الحقل "{field["label"]}".'
                )

    profile_fields = [
        field
        for category in categories
        for field in category["fields"]
        if field.get("image_display") == "profile"
    ]
    if len(profile_fields) > 1:
        raise ApplicationError("يمكن تحديد حقل صورة شخصية واحد فقط.")

    system_type_counts = {
        field_type: sum(
            field["type"] == field_type
            for category in categories
            for field in category["fields"]
        )
        for field_type in SYSTEM_FIELD_TYPES
    }
    if any(count > 1 for count in system_type_counts.values()):
        raise ApplicationError(
            "يمكن إضافة حقل واحد فقط من كل نوع من حقول بيانات السجل."
        )

    main_field_ids = {
        field["id"]
        for category in categories
        if category["kind"] == "main"
        for field in category["fields"]
        if field["type"] not in SYSTEM_FIELD_TYPES
    }

    for category in categories:
        anchor_field_id = category["anchor_field_id"]
        if anchor_field_id and anchor_field_id not in main_field_ids:
            raise ApplicationError(
                f'موضع الفئة "{category["label"]}" يجب أن يرتبط بحقل رئيسي.'
            )
        for field in category["fields"]:
            if field["type"] != "file":
                continue
            naming = field["file_naming"]
            if naming["mode"] == "template" and not naming["parts"]:
                raise ApplicationError(
                    f'أضف أجزاء تسمية الملف للحقل "{field["label"]}".'
                )
            for part in naming["parts"]:
                source_id = part["field_id"]
                if source_id not in field_lookup:
                    raise ApplicationError(
                        f'صيغة اسم الملف في "{field["label"]}" تشير إلى حقل محذوف.'
                    )
                source_category = category_lookup[field_category[source_id]]
                if (
                    source_category["kind"] != "main"
                    and source_category["id"] != category["id"]
                ):
                    raise ApplicationError(
                        f'صيغة اسم الملف في "{field["label"]}" لا يمكن أن تستخدم '
                        "حقلًا من فئة متكررة أخرى."
                    )
                if field_lookup[source_id]["type"] == "file":
                    raise ApplicationError(
                        f'صيغة اسم الملف في "{field["label"]}" لا يمكن أن تستخدم ملفًا آخر.'
                    )
                if field_lookup[source_id]["type"] in SYSTEM_FIELD_TYPES:
                    raise ApplicationError(
                        f'صيغة اسم الملف في "{field["label"]}" '
                        "لا تستخدم حقول بيانات السجل التقنية."
                    )

    for category in categories:
        for field in category["fields"]:
            validation = field.get("validation", {})
            compare_id = validation.get("compare_field_id")
            if compare_id:
                if compare_id not in field_lookup:
                    raise ApplicationError(
                        f'مقارنة الحقل "{field["label"]}" تشير إلى حقل محذوف.'
                    )
                compare_category = category_lookup[field_category[compare_id]]
                if field_lookup[compare_id]["type"] != field["type"]:
                    raise ApplicationError(
                        f'مقارنة الحقل "{field["label"]}" يجب أن تكون مع تاريخ من النوع نفسه.'
                    )
                if compare_category["kind"] != "main" and compare_category["id"] != category["id"]:
                    raise ApplicationError(
                        "لا يمكن مقارنة حقل بصف من فئة متكررة أخرى."
                    )

            raw_filter = field.pop("_option_filter_raw", None)
            if not raw_filter:
                continue
            if field["type"] not in {"select", "checkbox_group"}:
                raise ApplicationError(
                    f'تصفية الخيارات متاحة فقط للقائمة أو مجموعة الاختيارات في "{field["label"]}".'
                )
            if not isinstance(raw_filter, dict):
                raise ApplicationError(f'تصفية خيارات الحقل "{field["label"]}" غير صحيحة.')
            source_id = clean_text(raw_filter.get("source_field_id"))
            if source_id not in field_lookup:
                raise ApplicationError(f'مصدر تصفية الحقل "{field["label"]}" محذوف.')
            if source_id == field["id"]:
                raise ApplicationError("لا يمكن للحقل أن يصفّي خياراته بنفسه.")
            source = field_lookup[source_id]
            source_category = category_lookup[field_category[source_id]]
            if source["type"] not in {"select", "yes_no", "checkbox"}:
                raise ApplicationError(
                    f'الحقل المتحكم في "{field["label"]}" يجب أن يكون قائمة أو نعم/لا أو مربع اختيار.'
                )
            if source_category["kind"] != "main" and source_category["id"] != category["id"]:
                raise ApplicationError(
                    "تصفية فئة متكررة لا يمكن أن تعتمد على صف من فئة متكررة أخرى."
                )
            valid_source_tokens = (
                {"true", "false"}
                if source["type"] == "checkbox"
                else {option["id"] for option in source["options"]}
            )
            valid_target_ids = {option["id"] for option in field["options"]}
            raw_mappings = raw_filter.get("mappings", {})
            if not isinstance(raw_mappings, dict):
                raise ApplicationError(f'خريطة خيارات الحقل "{field["label"]}" غير صحيحة.')
            mappings: dict[str, list[str]] = {}
            for raw_token, raw_allowed in raw_mappings.items():
                token = clean_text(raw_token)
                if token not in valid_source_tokens:
                    continue
                if not isinstance(raw_allowed, list):
                    continue
                allowed: list[str] = []
                for option_id in raw_allowed:
                    option_id = clean_text(option_id)
                    if option_id in valid_target_ids and option_id not in allowed:
                        allowed.append(option_id)
                mappings[token] = allowed
            field["option_filter"] = {
                "source_field_id": source_id,
                "mappings": mappings,
                "unmatched": "none" if raw_filter.get("unmatched") == "none" else "all",
            }

    for category in categories:
        mode_source = related_person_mode_source_field(category)
        if mode_source:
            field_lookup[mode_source["id"]] = mode_source
            field_category[mode_source["id"]] = category["id"]

    raw_conditions = payload.get("conditions", [])

    if not isinstance(raw_conditions, list):
        raise ApplicationError(
            "قائمة الشروط غير صحيحة."
        )

    conditions: list[dict[str, Any]] = []
    condition_ids: set[str] = set()

    for raw_condition in raw_conditions:
        if not isinstance(raw_condition, dict):
            raise ApplicationError(
                "أحد شروط الظهور غير صحيح."
            )

        condition_id = _require_definition_id(
            raw_condition.get("id"),
            "cond",
            "معرّف الشرط",
        )

        if condition_id in condition_ids:
            raise ApplicationError(
                "يوجد معرّف شرط مكرر."
            )

        condition_ids.add(condition_id)

        target_type = raw_condition.get("target_type")
        target_id = clean_text(
            raw_condition.get("target_id")
        )
        source_field_id = clean_text(
            raw_condition.get("source_field_id")
        )

        if target_type not in {"category", "field"}:
            raise ApplicationError(
                "نوع هدف الشرط غير صالح."
            )

        if (
            target_type == "category"
            and target_id not in category_lookup
        ):
            raise ApplicationError(
                "هدف أحد الشروط هو فئة محذوفة."
            )

        if (
            target_type == "field"
            and target_id not in field_ids
        ):
            raise ApplicationError(
                "هدف أحد الشروط هو حقل محذوف."
            )

        if source_field_id not in field_lookup:
            raise ApplicationError(
                "مصدر أحد الشروط هو حقل محذوف."
            )

        if (
            source_field_id.startswith(RELATED_PERSON_MODE_SOURCE_PREFIX)
            and target_type != "field"
        ):
            raise ApplicationError(
                "حالة سجل الشخص المرتبط تتحكم في حقول بطاقته فقط."
            )

        if (
            target_type == "field"
            and target_id == source_field_id
        ):
            raise ApplicationError(
                "لا يمكن للحقل أن يتحكم في ظهوره بنفسه."
            )

        source_field = field_lookup[source_field_id]
        operator = raw_condition.get(
            "operator",
            "equals",
        )

        allowed_operators = condition_operators_for_field(
            source_field["type"]
        )

        if operator not in allowed_operators:
            raise ApplicationError(
                f'العملية المختارة لا تناسب نوع الحقل '
                f'"{source_field["label"]}".'
            )

        target_category_id = (
            target_id
            if target_type == "category"
            else field_category[target_id]
        )

        source_category = category_lookup[
            field_category[source_field_id]
        ]
        target_category = category_lookup[
            target_category_id
        ]

        if (
            source_category["kind"] != "main"
            and source_category["id"]
            != target_category["id"]
        ):
            raise ApplicationError(
                "شرط فئة متكررة لا يمكن أن يعتمد "
                "على صف من فئة متكررة أخرى."
            )

        default_group_id = stable_condition_group_id(
            target_type,
            target_id,
        )

        group_id = _require_definition_id(
            raw_condition.get("group_id")
            or default_group_id,
            "grp",
            "معرّف مجموعة الشرط",
        )

        value = normalize_condition_value(
            source_field,
            operator,
            raw_condition.get("value", ""),
        )

        conditions.append(
            {
                "id": condition_id,
                "group_id": group_id,
                "negate": bool(
                    raw_condition.get("negate", False)
                ),
                "target_type": target_type,
                "target_id": target_id,
                "source_field_id": source_field_id,
                "operator": operator,
                "value": value,
            }
        )
    revision = payload.get("revision", 0)
    try:
        revision = int(revision)
    except (TypeError, ValueError):
        revision = 0

    return {
        "schema_version": SCHEMA_VERSION,
        "revision": max(0, revision),
        "app": app,
        "categories": categories,
        "conditions": conditions,
    }


def schema_indexes(schema: dict[str, Any]) -> dict[str, Any]:
    categories = {category["id"]: category for category in schema["categories"]}
    fields: dict[str, dict[str, Any]] = {}
    field_categories: dict[str, str] = {}
    for category in schema["categories"]:
        for field in category["fields"]:
            fields[field["id"]] = field
            field_categories[field["id"]] = category["id"]
        mode_source = related_person_mode_source_field(category)
        if mode_source:
            fields[mode_source["id"]] = mode_source
            field_categories[mode_source["id"]] = category["id"]
    conditions_by_target: dict[tuple[str, str], list[dict[str, Any]]] = {}
    for condition in schema["conditions"]:
        key = (condition["target_type"], condition["target_id"])
        conditions_by_target.setdefault(key, []).append(condition)
    return {
        "categories": categories,
        "fields": fields,
        "field_categories": field_categories,
        "conditions_by_target": conditions_by_target,
    }


def read_schema_file() -> dict[str, Any]:
    if not SCHEMA_PATH.is_file():
        return default_schema()
    try:
        payload = json.loads(SCHEMA_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise ApplicationError(f"تعذّر قراءة ملف الإعدادات: {exc}") from exc
    try:
        version = int(payload.get("schema_version", 1))
    except (TypeError, ValueError):
        version = 0
    if version not in SUPPORTED_SCHEMA_VERSIONS:
        raise ApplicationError("إصدار ملف الإعدادات غير مدعوم.")
    schema = validate_schema(payload)
    if version != SCHEMA_VERSION:
        atomic_write_json(SCHEMA_PATH, schema)
    return schema


def _write_json_temp(path: Path, payload: dict[str, Any]) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(
        mode="w",
        encoding="utf-8",
        prefix=f".{path.stem}-",
        suffix=".json",
        dir=path.parent,
        delete=False,
    ) as temporary:
        json.dump(payload, temporary, ensure_ascii=False, indent=2)
        temporary.write("\n")
        return Path(temporary.name)


def atomic_write_json(path: Path, payload: dict[str, Any]) -> None:
    temporary = _write_json_temp(path, payload)
    try:
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def sanitize_sheet_name(value: str) -> str:
    name = INVALID_EXCEL_SHEET_PATTERN.sub(" ", value)
    name = " ".join(name.split()).strip("' ")
    return name or "فئة"


def related_sheet_names(schema: dict[str, Any]) -> dict[str, str]:
    used = {MAIN_SHEET.casefold(), META_SHEET.casefold()}
    result: dict[str, str] = {}
    for category in schema["categories"]:
        if category["kind"] != "repeatable":
            continue
        base = sanitize_sheet_name(category["label"])
        candidate = base[:31]
        if candidate.casefold() in used:
            suffix = f"-{category['id'][-6:]}"
            candidate = f"{base[:31 - len(suffix)]}{suffix}"
        while candidate.casefold() in used:
            candidate = f"{base[:22]}-{secrets.token_hex(4)}"[:31]
        used.add(candidate.casefold())
        result[category["id"]] = candidate
    return result


def main_fields(schema: dict[str, Any]) -> list[dict[str, Any]]:
    return [
        field
        for category in schema["categories"]
        if category["kind"] == "main"
        for field in category["fields"]
        if field["type"] not in SYSTEM_FIELD_TYPES
    ]


def file_fields(schema: dict[str, Any]) -> list[tuple[dict[str, Any], dict[str, Any]]]:
    return [
        (category, field)
        for category in schema["categories"]
        for field in category["fields"]
        if field["type"] == "file"
    ]


def technical_headers(worksheet) -> list[str]:
    row = next(
        worksheet.iter_rows(
            min_row=TECHNICAL_HEADER_ROW,
            max_row=TECHNICAL_HEADER_ROW,
            values_only=True,
        ),
        (),
    )
    headers = [clean_text(value) for value in row]
    while headers and not headers[-1]:
        headers.pop()
    return headers


def _sync_excel_field_labels_unlocked(
    schema: dict[str, Any],
) -> dict[str, Any]:
    """Import visible Excel label edits while preserving stable field IDs."""
    if not WORKBOOK_PATH.is_file() or not schema["categories"]:
        return schema
    try:
        workbook = load_workbook(WORKBOOK_PATH, read_only=True, data_only=False)
    except Exception as exc:
        raise ApplicationError(f"تعذّر قراءة ملف Excel: {exc}") from exc

    updated = copy.deepcopy(schema)
    changed = False
    try:
        category_sheets = related_sheet_names(updated)
        for category in updated["categories"]:
            if category["kind"] == "main":
                if MAIN_SHEET not in workbook.sheetnames:
                    continue
                worksheet = workbook[MAIN_SHEET]
            else:
                sheet_name = category_sheets[category["id"]]
                if sheet_name not in workbook.sheetnames:
                    continue
                worksheet = workbook[sheet_name]

            columns = {
                header: index
                for index, header in enumerate(
                    technical_headers(worksheet),
                    start=1,
                )
            }
            for field in category["fields"]:
                column = columns.get(field["id"])
                if column is None:
                    continue
                excel_label = clean_text(
                    worksheet.cell(VISIBLE_HEADER_ROW, column).value
                )
                if excel_label and excel_label != field["label"]:
                    field["label"] = excel_label
                    changed = True
    finally:
        workbook.close()

    if not changed:
        return schema

    updated["revision"] = schema["revision"] + 1
    updated = validate_schema(updated)
    atomic_write_json(SCHEMA_PATH, updated)
    return updated


def read_schema_with_excel_labels() -> dict[str, Any]:
    with WORKBOOK_LOCK:
        return _sync_excel_field_labels_unlocked(read_schema_file())


def _style_headers(worksheet, column_count: int) -> None:
    worksheet.row_dimensions[TECHNICAL_HEADER_ROW].hidden = True
    worksheet.row_dimensions[VISIBLE_HEADER_ROW].height = 28
    worksheet.freeze_panes = f"A{FIRST_DATA_ROW}"
    worksheet.sheet_view.showGridLines = False
    for column in range(1, column_count + 1):
        cell = worksheet.cell(VISIBLE_HEADER_ROW, column)
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        worksheet.column_dimensions[get_column_letter(column)].width = max(
            13, min(34, len(clean_text(cell.value)) + 4)
        )


def _add_field_validations(
    worksheet,
    field_columns: dict[str, int],
    fields: Iterable[dict[str, Any]],
) -> None:
    for field in fields:
        if field["id"] not in field_columns:
            continue
        options = option_labels(field, active_only=True)
        if field["type"] not in {"select", "yes_no"} or not options:
            continue
        escaped = ",".join(option.replace('"', '""') for option in options)
        if len(escaped) > 250:
            continue
        validation = DataValidation(
            type="list",
            formula1=f'"{escaped}"',
            allow_blank=True,
        )
        worksheet.add_data_validation(validation)
        column_letter = get_column_letter(field_columns[field["id"]])
        validation.add(f"{column_letter}{FIRST_DATA_ROW}:{column_letter}1048576")


def _new_workbook_structure(
    schema: dict[str, Any],
) -> tuple[Workbook, dict[str, Any]]:
    workbook = Workbook()
    main_sheet = workbook.active
    main_sheet.title = MAIN_SHEET

    fields = main_fields(schema)
    main_headers = [*MAIN_INTERNAL_HEADERS, *(field["id"] for field in fields)]
    main_labels = [
        "المعرّف الداخلي",
        "ID",
        "تاريخ الإنشاء",
        "تاريخ التعديل",
        "مؤرشف",
        "تاريخ الأرشفة",
        *(field["label"] for field in fields),
    ]
    main_sheet.append(main_headers)
    main_sheet.append(main_labels)
    _style_headers(main_sheet, len(main_headers))
    main_columns = {
        header: index for index, header in enumerate(main_headers, start=1)
    }
    _add_field_validations(main_sheet, main_columns, fields)

    sheets: dict[str, Any] = {"main": main_sheet, "related": {}}
    names = related_sheet_names(schema)
    for category in schema["categories"]:
        if category["kind"] != "repeatable":
            continue
        worksheet = workbook.create_sheet(names[category["id"]])
        headers = [
            *RELATED_INTERNAL_HEADERS,
            RELATED_LINK_HEADER,
            *(marker["id"] for marker in category.get("row_markers", [])),
            *(field["id"] for field in category["fields"]),
        ]
        labels = [
            "معرّف الصف الداخلي",
            "معرّف السجل الداخلي",
            "ID",
            "minor_id",
            "تاريخ الإنشاء",
            "تاريخ التعديل",
            "ID الشخص المرتبط",
            *(marker["display_text"] for marker in category.get("row_markers", [])),
            *(field["label"] for field in category["fields"]),
        ]
        worksheet.append(headers)
        worksheet.append(labels)
        _style_headers(worksheet, len(headers))
        columns = {header: index for index, header in enumerate(headers, start=1)}
        _add_field_validations(worksheet, columns, category["fields"])
        sheets["related"][category["id"]] = worksheet

    meta = workbook.create_sheet(META_SHEET)
    meta.sheet_state = "hidden"
    meta.append(["schema_version", SCHEMA_VERSION])
    meta.append(["schema_revision", schema["revision"]])
    meta.append(["category_id", "sheet_name"])
    for category_id, sheet_name in names.items():
        meta.append([category_id, sheet_name])
    return workbook, sheets


def _apply_date_formats(
    worksheet,
    row_index: int,
    field_columns: dict[str, int],
    fields: Iterable[dict[str, Any]],
) -> None:
    for field in fields:
        if field["type"] == "date_gregorian" and field["id"] in field_columns:
            worksheet.cell(
                row=row_index, column=field_columns[field["id"]]
            ).number_format = EXCEL_DATE_FORMAT


def write_dataset_workbook(
    schema: dict[str, Any],
    records: list[dict[str, Any]],
    destination: Path,
) -> None:
    workbook, sheets = _new_workbook_structure(schema)
    try:
        main_sheet = sheets["main"]
        main_headers = technical_headers(main_sheet)
        main_columns = {
            header: index for index, header in enumerate(main_headers, start=1)
        }
        fields = main_fields(schema)

        for record in records:
            row = [
                record["_record_id"],
                record["record_code"],
                record["created_at"],
                record["updated_at"],
                "نعم" if record.get("archived") else "لا",
                record.get("archived_at", ""),
                *[
                    excel_value(record.get("values", {}).get(field["id"], ""), field)
                    for field in fields
                ],
            ]
            main_sheet.append(row)
            _apply_date_formats(
                main_sheet, main_sheet.max_row, main_columns, fields
            )

        for category in schema["categories"]:
            if category["kind"] != "repeatable":
                continue
            worksheet = sheets["related"][category["id"]]
            headers = technical_headers(worksheet)
            columns = {
                header: index for index, header in enumerate(headers, start=1)
            }
            for record in records:
                rows = record.get("related", {}).get(category["id"], [])
                for sequence, child in enumerate(rows, start=1):
                    row = [
                        child["_child_id"],
                        record["_record_id"],
                        record["record_code"],
                        sequence,
                        child["created_at"],
                        child["updated_at"],
                        child.get("linked_record_code", ""),
                        *[
                            "نعم" if child.get("markers", {}).get(marker["id"]) else "لا"
                            for marker in category.get("row_markers", [])
                        ],
                        *[
                            excel_value(
                                child.get("values", {}).get(field["id"], ""),
                                field,
                            )
                            for field in category["fields"]
                        ],
                    ]
                    worksheet.append(row)
                    _apply_date_formats(
                        worksheet,
                        worksheet.max_row,
                        columns,
                        category["fields"],
                    )

        for worksheet in workbook.worksheets:
            if worksheet.title == META_SHEET:
                continue
            last_row = max(VISIBLE_HEADER_ROW, worksheet.max_row)
            last_column = max(1, worksheet.max_column)
            worksheet.auto_filter.ref = (
                f"A{VISIBLE_HEADER_ROW}:"
                f"{get_column_letter(last_column)}{last_row}"
            )
        destination.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(destination)
    finally:
        workbook.close()


def atomic_write_workbook(
    schema: dict[str, Any], records: list[dict[str, Any]]
) -> None:
    WORKBOOK_PATH.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(
        prefix=".database-",
        suffix=".xlsx",
        dir=WORKBOOK_PATH.parent,
        delete=False,
    ) as temporary:
        temporary_path = Path(temporary.name)
    try:
        started = time.perf_counter()
        write_dataset_workbook(schema, records, temporary_path)
        os.replace(temporary_path, WORKBOOK_PATH)
        invalidate_dataset_cache()
        LOGGER.info(
            "Workbook written: %.3f seconds; records=%d",
            time.perf_counter() - started,
            len(records),
        )
    finally:
        temporary_path.unlink(missing_ok=True)


def _temporary_copy(path: Path, prefix: str) -> Path:
    with tempfile.NamedTemporaryFile(
        prefix=prefix,
        suffix=path.suffix,
        dir=path.parent,
        delete=False,
    ) as temporary:
        temporary_path = Path(temporary.name)
    try:
        shutil.copy2(path, temporary_path)
    except Exception:
        temporary_path.unlink(missing_ok=True)
        raise
    return temporary_path


def _replace_storage_pair(workbook_temp: Path, schema_temp: Path) -> None:
    """Commit workbook and schema together, rolling the workbook back on failure."""
    workbook_backup = _temporary_copy(
        WORKBOOK_PATH,
        ".database-rollback-",
    )
    workbook_replaced = False
    try:
        os.replace(workbook_temp, WORKBOOK_PATH)
        workbook_replaced = True
        os.replace(schema_temp, SCHEMA_PATH)
        invalidate_dataset_cache()
    except Exception as exc:
        if workbook_replaced:
            try:
                os.replace(workbook_backup, WORKBOOK_PATH)
                workbook_backup = None
            except Exception as recovery_exc:
                raise ApplicationError(
                    "تعذّر إكمال تحديث التصميم وتعذّرت الاستعادة التلقائية. "
                    "لا تفتح التطبيق مجددًا قبل استعادة آخر نسخة احتياطية."
                ) from recovery_exc
        raise exc
    finally:
        workbook_temp.unlink(missing_ok=True)
        schema_temp.unlink(missing_ok=True)
        if workbook_backup is not None:
            workbook_backup.unlink(missing_ok=True)


def ensure_storage() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    schema_created = False
    if not SCHEMA_PATH.is_file():
        atomic_write_json(SCHEMA_PATH, default_schema())
        schema_created = True
    schema = read_schema_file()
    if not WORKBOOK_PATH.is_file() or schema_created:
        atomic_write_workbook(schema, [])
    attachments_directory().mkdir(parents=True, exist_ok=True)


def _row_value(
    row: tuple[Any, ...],
    columns: dict[str, int],
    header: str,
) -> Any:
    """Read from an iter_rows tuple using a one-based header map."""
    column = columns.get(header)
    if column is None or column > len(row):
        return None
    return row[column - 1]


def read_excel_field_value(value: Any, field: dict[str, Any]) -> Any:
    raw = json_value(value)
    if field["type"] == "checkbox":
        return clean_text(raw).casefold() in {"نعم", "true", "1", "yes"}
    if field["type"] == "checkbox_group":
        return [clean_text(item) for item in parse_checkbox_group_value(raw) if clean_text(item)]
    return raw


def excel_boolean(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return clean_text(value).casefold() in {"نعم", "true", "1", "yes"}




def read_dataset_unlocked(schema: dict[str, Any]) -> list[dict[str, Any]]:
    if not WORKBOOK_PATH.is_file():
        return []
    try:
        workbook = load_workbook(WORKBOOK_PATH, read_only=True, data_only=False)
    except Exception as exc:
        raise ApplicationError(f"تعذّر قراءة ملف Excel: {exc}") from exc

    try:
        if MAIN_SHEET not in workbook.sheetnames:
            raise ApplicationError(
                "ملف Excel لا يتوافق مع إصدار التطبيق العام الحالي."
            )
        main_sheet = workbook[MAIN_SHEET]
        headers = technical_headers(main_sheet)
        missing = [header for header in MAIN_REQUIRED_INTERNAL_HEADERS if header not in headers]
        if missing:
            raise ApplicationError(
                "ملف Excel لا يحتوي على الأعمدة الداخلية المطلوبة."
            )
        columns = {header: index for index, header in enumerate(headers, start=1)}
        fields = main_fields(schema)
        records: list[dict[str, Any]] = []
        by_internal_id: dict[str, dict[str, Any]] = {}
        seen_record_codes: set[str] = set()

        for row in main_sheet.iter_rows(
            min_row=FIRST_DATA_ROW,
            values_only=True,
        ):
            record_id = clean_text(_row_value(row, columns, "_record_id"))
            record_code = clean_text(_row_value(row, columns, "record_code"))
            if not record_id and not record_code:
                continue
            if not INTERNAL_ID_PATTERN.fullmatch(record_id):
                raise ApplicationError("يوجد معرّف سجل داخلي تالف في Excel.")
            if not PERSON_CODE_PATTERN.fullmatch(record_code):
                raise ApplicationError("يوجد معرّف سجل ظاهر تالف في Excel.")
            if record_id in by_internal_id:
                raise ApplicationError(
                    "يوجد معرّف سجل داخلي مكرر في Excel. "
                    "استعد نسخة سليمة قبل متابعة التعديل."
                )
            if record_code in seen_record_codes:
                raise ApplicationError(
                    f'معرّف السجل الظاهر "{record_code}" مكرر في Excel.'
                )
            record = {
                "_record_id": record_id,
                "record_code": record_code,
                "created_at": clean_text(_row_value(row, columns, "created_at")),
                "updated_at": clean_text(_row_value(row, columns, "updated_at")),
                "archived": excel_boolean(_row_value(row, columns, "_archived")),
                "archived_at": clean_text(_row_value(row, columns, "_archived_at")),
                "values": {
                    field["id"]: read_excel_field_value(
                        _row_value(row, columns, field["id"]), field
                    )
                    for field in fields
                },
                "related": {},
            }
            records.append(record)
            by_internal_id[record_id] = record
            seen_record_codes.add(record_code)

        names = related_sheet_names(schema)
        seen_child_ids: set[str] = set()
        for category in schema["categories"]:
            if category["kind"] != "repeatable":
                continue
            for record in records:
                record["related"][category["id"]] = []
            sheet_name = names[category["id"]]
            if sheet_name not in workbook.sheetnames:
                continue
            worksheet = workbook[sheet_name]
            related_headers = technical_headers(worksheet)
            related_columns = {
                header: index
                for index, header in enumerate(related_headers, start=1)
            }
            missing = [
                header
                for header in RELATED_INTERNAL_HEADERS
                if header not in related_columns
            ]
            if missing:
                raise ApplicationError(
                    f'ورقة "{sheet_name}" لا تحتوي على الأعمدة الداخلية المطلوبة.'
                )
            for row in worksheet.iter_rows(
                min_row=FIRST_DATA_ROW,
                values_only=True,
            ):
                record_id = clean_text(
                    _row_value(row, related_columns, "_record_id")
                )
                child_id = clean_text(
                    _row_value(row, related_columns, "_child_id")
                )
                related_code = clean_text(
                    _row_value(row, related_columns, "record_code")
                )
                if not record_id and not child_id and not related_code:
                    continue
                if record_id not in by_internal_id:
                    raise ApplicationError(
                        f'يوجد صف في ورقة "{sheet_name}" لا يعود إلى سجل معروف.'
                    )
                if related_code != by_internal_id[record_id]["record_code"]:
                    raise ApplicationError(
                        f'يوجد صف في ورقة "{sheet_name}" يحمل معرّف سجل غير متطابق.'
                    )
                if not INTERNAL_ID_PATTERN.fullmatch(child_id):
                    raise ApplicationError(
                        f'يوجد معرّف صف داخلي تالف في ورقة "{sheet_name}".'
                    )
                if child_id in seen_child_ids:
                    raise ApplicationError(
                        f'يوجد معرّف صف داخلي مكرر في ورقة "{sheet_name}".'
                    )
                minor_id = _row_value(row, related_columns, "minor_id")
                try:
                    minor_number = int(minor_id)
                except (TypeError, ValueError) as exc:
                    raise ApplicationError(
                        f'يوجد minor_id غير صالح في ورقة "{sheet_name}".'
                    ) from exc
                if minor_number < 1:
                    raise ApplicationError(
                        f'يوجد minor_id غير صالح في ورقة "{sheet_name}".'
                    )
                child = {
                    "_child_id": child_id,
                    "minor_id": minor_number,
                    "created_at": clean_text(
                        _row_value(row, related_columns, "created_at")
                    ),
                    "updated_at": clean_text(
                        _row_value(row, related_columns, "updated_at")
                    ),
                    "linked_record_code": clean_text(
                        _row_value(row, related_columns, RELATED_LINK_HEADER)
                    ),
                    "markers": {
                        marker["id"]: excel_boolean(
                            _row_value(row, related_columns, marker["id"])
                        )
                        for marker in category.get("row_markers", [])
                    },
                    "values": {
                        field["id"]: read_excel_field_value(
                            _row_value(row, related_columns, field["id"]), field
                        )
                        for field in category["fields"]
                    },
                }
                by_internal_id[record_id]["related"][category["id"]].append(child)
                seen_child_ids.add(child_id)

            for record in records:
                record["related"][category["id"]].sort(
                    key=lambda child: (
                        int(child.get("minor_id") or 0),
                        child["_child_id"],
                    )
                )
        return records
    finally:
        workbook.close()


def _schema_cache_signature(schema: dict[str, Any]) -> str:
    payload = json.dumps(
        schema,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("utf-8")
    return hashlib.sha256(payload).hexdigest()


def _workbook_cache_signature() -> tuple[int, int] | None:
    try:
        status = WORKBOOK_PATH.stat()
    except FileNotFoundError:
        return None
    return status.st_mtime_ns, status.st_size


def invalidate_dataset_cache() -> None:
    global _DATASET_SNAPSHOT
    _DATASET_SNAPSHOT = None


def _search_index_value(field: dict[str, Any], value: Any) -> Any:
    field_type = field["type"]
    if field_type == "checkbox":
        return excel_boolean(value)
    if field_type == "checkbox_group":
        return frozenset(
            normalize_search_text(item)
            for item in parse_checkbox_group_value(value)
        )
    if field_type == "number":
        try:
            return float(value)
        except (TypeError, ValueError):
            return None
    return normalize_search_text(value)


def _snapshot_record_copy(
    schema: dict[str, Any], record: dict[str, Any]
) -> dict[str, Any]:
    snapshot_record = copy.deepcopy(record)
    for category in schema["categories"]:
        if category["kind"] == "main":
            values = snapshot_record.setdefault("values", {})
            for field in category["fields"]:
                if field["type"] in SYSTEM_FIELD_TYPES:
                    values.pop(field["id"], None)
                    continue
                values[field["id"]] = read_excel_field_value(
                    values.get(field["id"], ""), field
                )
            continue
        for row in snapshot_record.setdefault("related", {}).setdefault(
            category["id"], []
        ):
            values = row.setdefault("values", {})
            for field in category["fields"]:
                values[field["id"]] = read_excel_field_value(
                    values.get(field["id"], ""), field
                )
    return snapshot_record


def _build_dataset_snapshot(
    schema: dict[str, Any],
    records: Iterable[dict[str, Any]],
    *,
    workbook_signature: tuple[int, int] | None = None,
) -> DatasetSnapshot:
    record_tuple = tuple(
        _snapshot_record_copy(schema, record) for record in records
    )
    records_by_code: dict[str, dict[str, Any]] = {}
    records_by_id: dict[str, dict[str, Any]] = {}
    record_positions: dict[str, int] = {}
    unique_mutable: dict[str, dict[str, set[str]]] = {}
    search_main: dict[str, dict[str, Any]] = {}
    search_related: dict[str, dict[str, tuple[dict[str, Any], ...]]] = {}

    searchable_fields = {
        field["id"]: field
        for category in schema["categories"]
        for field in category["fields"]
        if field["type"] != "file" and field["type"] not in SYSTEM_FIELD_TYPES
    }
    unique_fields = {
        field["id"]: (category, field)
        for category in schema["categories"]
        for field in category["fields"]
        if field.get("unique")
    }

    for position, record in enumerate(record_tuple):
        record_id = record["_record_id"]
        record_code = record["record_code"]
        records_by_code[record_code] = record
        records_by_id[record_id] = record
        record_positions[record_id] = position

        main_values: dict[str, Any] = {}
        related_values: dict[str, tuple[dict[str, Any], ...]] = {}
        for category in schema["categories"]:
            if category["kind"] == "main":
                for field in category["fields"]:
                    if field["id"] in searchable_fields:
                        main_values[field["id"]] = _search_index_value(
                            field, record.get("values", {}).get(field["id"], "")
                        )
            else:
                indexed_rows: list[dict[str, Any]] = []
                for row in record.get("related", {}).get(category["id"], []):
                    indexed_rows.append(
                        {
                            field["id"]: _search_index_value(
                                field, row.get("values", {}).get(field["id"], "")
                            )
                            for field in category["fields"]
                            if field["id"] in searchable_fields
                        }
                    )
                related_values[category["id"]] = tuple(indexed_rows)
        search_main[record_id] = main_values
        search_related[record_id] = related_values

        for field_id, (category, field) in unique_fields.items():
            for value in unique_field_values_in_record(record, category, field):
                if value == "" or value is None or value == []:
                    continue
                token = canonical_unique_value(value)
                unique_mutable.setdefault(field_id, {}).setdefault(
                    token, set()
                ).add(record_id)

    unique_indexes = {
        field_id: {
            token: frozenset(owners)
            for token, owners in values.items()
        }
        for field_id, values in unique_mutable.items()
    }

    return DatasetSnapshot(
        workbook_path=str(WORKBOOK_PATH.resolve()),
        workbook_signature=(
            _workbook_cache_signature()
            if workbook_signature is None
            else workbook_signature
        ),
        schema_signature=_schema_cache_signature(schema),
        records=record_tuple,
        records_by_code=records_by_code,
        records_by_id=records_by_id,
        record_positions=record_positions,
        unique_indexes=unique_indexes,
        search_main=search_main,
        search_related=search_related,
    )


def _dataset_snapshot_unlocked(
    schema: dict[str, Any],
    *,
    force_reload: bool = False,
) -> DatasetSnapshot:
    global _DATASET_SNAPSHOT
    workbook_path = str(WORKBOOK_PATH.resolve())
    workbook_signature = _workbook_cache_signature()
    schema_signature = _schema_cache_signature(schema)

    if (
        not force_reload
        and _DATASET_SNAPSHOT is not None
        and _DATASET_SNAPSHOT.workbook_path == workbook_path
        and _DATASET_SNAPSHOT.workbook_signature == workbook_signature
        and _DATASET_SNAPSHOT.schema_signature == schema_signature
    ):
        return _DATASET_SNAPSHOT

    started = time.perf_counter()
    records = read_dataset_unlocked(schema)
    _DATASET_SNAPSHOT = _build_dataset_snapshot(
        schema, records, workbook_signature=workbook_signature
    )
    LOGGER.info(
        "Workbook loaded and indexed: %.3f seconds; records=%d",
        time.perf_counter() - started,
        len(records),
    )
    return _DATASET_SNAPSHOT


def _publish_dataset_snapshot(
    schema: dict[str, Any], records: Iterable[dict[str, Any]]
) -> DatasetSnapshot:
    global _DATASET_SNAPSHOT
    _DATASET_SNAPSHOT = _build_dataset_snapshot(schema, records)
    return _DATASET_SNAPSHOT


def _safe_publish_dataset_snapshot(
    schema: dict[str, Any], records: Iterable[dict[str, Any]]
) -> None:
    try:
        _publish_dataset_snapshot(schema, records)
    except Exception:
        # The workbook commit has already succeeded. A cache failure must never
        # turn a successful save into an apparent failure or remove attachments.
        invalidate_dataset_cache()
        LOGGER.exception("Failed to rebuild the in-memory dataset snapshot")


def read_dataset(schema: dict[str, Any]) -> list[dict[str, Any]]:
    with WORKBOOK_LOCK:
        # Public callers receive an isolated copy; only internal read paths use
        # the immutable published snapshot directly.
        return copy.deepcopy(list(_dataset_snapshot_unlocked(schema).records))


def record_count(schema: dict[str, Any] | None = None) -> int:
    schema = schema or read_schema_file()
    with WORKBOOK_LOCK:
        return len(_dataset_snapshot_unlocked(schema).records)


def _record_file_paths(
    schema: dict[str, Any], record: dict[str, Any]
) -> set[str]:
    paths: set[str] = set()
    for category, field in file_fields(schema):
        if category["kind"] == "main":
            raw_values = [record.get("values", {}).get(field["id"], "")]
        else:
            raw_values = [
                child.get("values", {}).get(field["id"], "")
                for child in record.get("related", {}).get(category["id"], [])
            ]
        for raw in raw_values:
            relative = attachment_relative_path(raw)
            if relative:
                paths.add(relative)
    return paths


def migrate_field_value(
    value: Any,
    old_field: dict[str, Any] | None,
    new_field: dict[str, Any],
) -> Any:
    if old_field is None:
        return value
    old_type = old_field["type"]
    new_type = new_field["type"]
    if (old_type == "file") != (new_type == "file"):
        return ""
    if old_type == new_type and new_type in {"select", "yes_no"}:
        old_option = option_by_value(old_field, value)
        if old_option:
            new_option = option_by_id(new_field, old_option["id"])
            return new_option["label"] if new_option else ""
    if old_type == new_type == "checkbox_group":
        migrated: list[str] = []
        for item in parse_checkbox_group_value(value):
            old_option = option_by_value(old_field, item)
            if not old_option:
                continue
            new_option = option_by_id(new_field, old_option["id"])
            if new_option and new_option["label"] not in migrated:
                migrated.append(new_option["label"])
        return migrated
    if new_type == "checkbox" and old_type != "checkbox":
        return False
    if old_type == "checkbox" and new_type != "checkbox":
        return ""
    if new_type == "checkbox_group" and old_type != "checkbox_group":
        return []
    if old_type == "checkbox_group" and new_type != "checkbox_group":
        return ""
    return value




def migrate_records_to_schema(
    old_schema: dict[str, Any],
    new_schema: dict[str, Any],
    records: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    old_indexes = schema_indexes(old_schema)
    new_indexes = schema_indexes(new_schema)

    for field_id, old_category_id in old_indexes["field_categories"].items():
        if field_id not in new_indexes["field_categories"]:
            continue
        new_category_id = new_indexes["field_categories"][field_id]
        old_kind = old_indexes["categories"][old_category_id]["kind"]
        new_kind = new_indexes["categories"][new_category_id]["kind"]
        if old_category_id != new_category_id or old_kind != new_kind:
            if records:
                raise ApplicationError(
                    "لا يمكن نقل حقل إلى فئة أخرى بعد وجود سجلات. "
                    "أنشئ حقلًا جديدًا بدلًا منه."
                )

    for category_id, old_category in old_indexes["categories"].items():
        if category_id in new_indexes["categories"]:
            new_category = new_indexes["categories"][category_id]
            if old_category["kind"] != new_category["kind"] and records:
                raise ApplicationError(
                    f'لا يمكن تغيير نوع الفئة "{old_category["label"]}" بعد وجود سجلات.'
                )

    old_fields = old_indexes["fields"]
    new_fields = new_indexes["fields"]
    new_main_ids = {
        field["id"]
        for category in new_schema["categories"]
        if category["kind"] == "main"
        for field in category["fields"]
        if field["type"] not in SYSTEM_FIELD_TYPES
    }
    new_related_ids = {
        category["id"]: {field["id"] for field in category["fields"]}
        for category in new_schema["categories"]
        if category["kind"] == "repeatable"
    }

    migrated = copy.deepcopy(records)
    for record in migrated:
        new_values: dict[str, Any] = {}
        for field_id in new_main_ids:
            value = record.get("values", {}).get(field_id, "")
            value = migrate_field_value(
                value,
                old_fields.get(field_id),
                new_fields[field_id],
            )
            new_values[field_id] = value
        record["values"] = new_values

        new_related: dict[str, list[dict[str, Any]]] = {}
        for category_id, field_id_set in new_related_ids.items():
            rows = record.get("related", {}).get(category_id, [])
            new_rows = []
            for row in rows:
                values = {}
                for field_id in field_id_set:
                    value = row.get("values", {}).get(field_id, "")
                    value = migrate_field_value(
                        value,
                        old_fields.get(field_id),
                        new_fields[field_id],
                    )
                    values[field_id] = value
                row["values"] = values
                valid_marker_ids = {
                    marker["id"]
                    for marker in new_indexes["categories"][category_id].get("row_markers", [])
                }
                row["markers"] = {
                    marker_id: bool(selected)
                    for marker_id, selected in row.get("markers", {}).items()
                    if marker_id in valid_marker_ids
                }
                if not new_indexes["categories"][category_id].get(
                    "related_person_enabled"
                ):
                    row["linked_record_code"] = ""
                new_rows.append(row)
            new_related[category_id] = new_rows
        record["related"] = new_related
    return migrated


def schema_response(schema: dict[str, Any]) -> dict[str, Any]:
    response = copy.deepcopy(schema)
    with WORKBOOK_LOCK:
        records = _dataset_snapshot_unlocked(schema).records
    access = builder_access_response()
    response["developer_mode"] = access["unlocked"]
    response["builder_access"] = access
    response["stats"] = {
        "record_count": len(records),
        "field_count": sum(
            len(category["fields"]) for category in schema["categories"]
        ),
        "category_count": len(schema["categories"]),
    }
    response["archive_stats"] = {
        "active_record_count": sum(not record.get("archived") for record in records),
        "archived_record_count": sum(bool(record.get("archived")) for record in records),
    }
    return response


def schema_change_is_destructive(
    old_schema: dict[str, Any], new_schema: dict[str, Any]
) -> bool:
    old_indexes = schema_indexes(old_schema)
    new_indexes = schema_indexes(new_schema)
    if set(old_indexes["categories"]) - set(new_indexes["categories"]):
        return True
    if set(old_indexes["fields"]) - set(new_indexes["fields"]):
        return True
    for category_id, old_category in old_indexes["categories"].items():
        new_category = new_indexes["categories"].get(category_id)
        if not new_category:
            continue
        if old_category["kind"] != new_category["kind"]:
            return True
        old_markers = {marker["id"] for marker in old_category.get("row_markers", [])}
        new_markers = {marker["id"] for marker in new_category.get("row_markers", [])}
        if old_markers - new_markers:
            return True
    for field_id, old_field in old_indexes["fields"].items():
        new_field = new_indexes["fields"].get(field_id)
        if not new_field:
            continue
        if old_field["type"] != new_field["type"]:
            return True
        old_options = {option["id"] for option in old_field.get("options", [])}
        new_options = {option["id"] for option in new_field.get("options", [])}
        if old_options - new_options:
            return True
    return False




def save_schema(payload: Any) -> dict[str, Any]:
    new_schema = validate_schema(payload)
    expected_revision = new_schema["revision"]
    with WORKBOOK_LOCK:
        current = _sync_excel_field_labels_unlocked(read_schema_file())
        if expected_revision != current["revision"]:
            raise ApplicationError(
                "تغيّر الإعداد في نافذة أخرى. أعد تحميله قبل الحفظ."
            )
        records = copy.deepcopy(_dataset_snapshot_unlocked(current).records)
        automatic_backup = None
        if records and schema_change_is_destructive(current, new_schema):
            automatic_backup = create_backup(automatic=True)
        old_paths = {
            path
            for record in records
            for path in _record_file_paths(current, record)
        }
        new_schema["revision"] = current["revision"] + 1
        migrated = migrate_records_to_schema(current, new_schema, records)
        new_paths = {
            path
            for record in migrated
            for path in _record_file_paths(new_schema, record)
        }

        workbook_temp: Path | None = None
        schema_temp: Path | None = None
        try:
            with tempfile.NamedTemporaryFile(
                prefix=".database-schema-",
                suffix=".xlsx",
                dir=DATA_DIR,
                delete=False,
            ) as temporary:
                workbook_temp = Path(temporary.name)
            write_dataset_workbook(new_schema, migrated, workbook_temp)
            schema_temp = _write_json_temp(SCHEMA_PATH, new_schema)
            _replace_storage_pair(workbook_temp, schema_temp)
            workbook_temp = None
            schema_temp = None
        except PermissionError as exc:
            raise ApplicationError(
                "تعذّر تحديث الإعداد. أغلق database.xlsx في Excel ثم حاول مجددًا."
            ) from exc
        finally:
            if workbook_temp:
                workbook_temp.unlink(missing_ok=True)
            if schema_temp:
                schema_temp.unlink(missing_ok=True)
        remove_attachment_files(old_paths - new_paths)
        _safe_publish_dataset_snapshot(new_schema, migrated)
    response = schema_response(new_schema)
    if automatic_backup:
        response["automatic_backup"] = automatic_backup
    return response


def condition_value_is_empty(
    value: Any,
    source_field: dict[str, Any],
) -> bool:
    field_type = source_field["type"]

    if field_type == "checkbox_group":
        return not parse_checkbox_group_value(value)

    if field_type == "file":
        if isinstance(value, dict):
            return not (
                value.get("stored_path")
                or value.get("upload")
            )

        return not clean_text(value)

    if value is None:
        return True

    return clean_text(json_value(value)) == ""


def condition_matches(
    condition: dict[str, Any],
    main_values: dict[str, Any],
    row_values: dict[str, Any] | None,
    indexes: dict[str, Any],
    target_category_id: str,
) -> bool:
    source_id = condition["source_field_id"]
    source_category_id = indexes[
        "field_categories"
    ][source_id]
    source_category = indexes[
        "categories"
    ][source_category_id]

    if (
        source_category["kind"] == "repeatable"
        and source_category_id
        == target_category_id
    ):
        value = (row_values or {}).get(
            source_id,
            "",
        )
    else:
        value = main_values.get(
            source_id,
            "",
        )

    source_field = indexes["fields"][source_id]
    field_type = source_field["type"]
    operator = condition["operator"]
    expected = condition.get("value", "")

    empty = condition_value_is_empty(
        value,
        source_field,
    )

    if operator == "empty":
        result = empty

    elif operator == "not_empty":
        result = not empty

    elif empty:
        result = False

    elif field_type == "checkbox":
        actual = (
            "true"
            if excel_boolean(value)
            else "false"
        )

        result = (
            actual == expected
            if operator == "equals"
            else actual != expected
        )

    elif field_type in {"select", "yes_no"}:
        option = option_by_value(
            source_field,
            value,
        )
        actual = option["id"] if option else ""

        result = (
            actual == expected
            if operator == "equals"
            else actual != expected
        )

    elif field_type == "checkbox_group":
        actual_ids: set[str] = set()

        for item in parse_checkbox_group_value(value):
            option = option_by_value(
                source_field,
                item,
            )

            if option:
                actual_ids.add(option["id"])

        if operator == "contains":
            result = expected in actual_ids
        else:
            result = expected not in actual_ids

    elif field_type == "number":
        try:
            actual_number = float(value)
            expected_number = float(expected)
        except (TypeError, ValueError):
            result = False
        else:
            result = {
                "equals":
                    actual_number == expected_number,
                "not_equals":
                    actual_number != expected_number,
                "greater_than":
                    actual_number > expected_number,
                "greater_or_equal":
                    actual_number >= expected_number,
                "less_than":
                    actual_number < expected_number,
                "less_or_equal":
                    actual_number <= expected_number,
            }.get(operator, False)

    elif field_type.startswith("date_"):
        actual_date = clean_text(
            json_value(value)
        )

        result = {
            "equals":
                actual_date == expected,
            "not_equals":
                actual_date != expected,
            "before":
                actual_date < expected,
            "after":
                actual_date > expected,
            "on_or_before":
                actual_date <= expected,
            "on_or_after":
                actual_date >= expected,
        }.get(operator, False)

    else:
        actual_text = normalize_search_text(value)
        expected_text = normalize_search_text(
            expected
        )

        result = {
            "equals":
                actual_text == expected_text,
            "not_equals":
                actual_text != expected_text,
            "contains":
                expected_text in actual_text,
            "not_contains":
                expected_text not in actual_text,
        }.get(operator, False)

    if condition.get("negate", False):
        return not result

    return result


def target_visible(
    target_type: str,
    target_id: str,
    main_values: dict[str, Any],
    row_values: dict[str, Any] | None,
    indexes: dict[str, Any],
) -> bool:
    rules = indexes[
        "conditions_by_target"
    ].get(
        (target_type, target_id),
        [],
    )

    if not rules:
        return True

    target_category_id = (
        target_id
        if target_type == "category"
        else indexes["field_categories"][target_id]
    )

    groups: dict[
        str,
        list[dict[str, Any]],
    ] = {}

    for rule in rules:
        group_id = (
            rule.get("group_id")
            or stable_condition_group_id(
                target_type,
                target_id,
            )
        )

        groups.setdefault(
            group_id,
            [],
        ).append(rule)

    # OR between groups, AND inside each group.
    return any(
        all(
            condition_matches(
                rule,
                main_values,
                row_values,
                indexes,
                target_category_id,
            )
            for rule in group_rules
        )
        for group_rules in groups.values()
    )

def field_source_value(
    source_field_id: str,
    target_category_id: str,
    main_values: dict[str, Any],
    row_values: dict[str, Any] | None,
    indexes: dict[str, Any],
) -> Any:
    source_category_id = indexes["field_categories"][source_field_id]
    source_category = indexes["categories"][source_category_id]
    if source_category["kind"] == "repeatable" and source_category_id == target_category_id:
        return (row_values or {}).get(source_field_id, "")
    return main_values.get(source_field_id, "")


def allowed_option_ids(
    field: dict[str, Any],
    main_values: dict[str, Any],
    row_values: dict[str, Any] | None,
    indexes: dict[str, Any],
) -> set[str]:
    all_ids = {
        option["id"] for option in field.get("options", []) if option.get("active", True)
    }
    option_filter = field.get("option_filter")
    if not option_filter:
        return all_ids
    target_category_id = indexes["field_categories"][field["id"]]
    source_id = option_filter["source_field_id"]
    source = indexes["fields"][source_id]
    source_value = field_source_value(
        source_id,
        target_category_id,
        main_values,
        row_values,
        indexes,
    )
    token = option_token(source, source_value)
    mappings = option_filter.get("mappings", {})
    if token in mappings:
        return set(mappings[token]) & all_ids
    return set() if option_filter.get("unmatched") == "none" else all_ids


def validate_dependent_option_value(
    field: dict[str, Any],
    normalized: Any,
    main_values: dict[str, Any],
    row_values: dict[str, Any] | None,
    indexes: dict[str, Any],
) -> None:
    if field["type"] not in {"select", "checkbox_group"} or normalized is None or normalized == "":
        return
    allowed = allowed_option_ids(field, main_values, row_values, indexes)
    values = normalized if field["type"] == "checkbox_group" else [normalized]
    for value in values:
        option = option_by_value(field, value)
        if option is None or option["id"] not in allowed:
            raise ApplicationError(
                f'قيمة الحقل "{field["label"]}" غير متاحة وفق الحقل المتحكم.'
            )


def validate_marker_rules(category: dict[str, Any], rows: list[dict[str, Any]]) -> None:
    for marker in category.get("row_markers", []):
        selected = sum(bool(row.get("markers", {}).get(marker["id"])) for row in rows)
        rule = marker["rule"]
        if rule == "at_most_one" and selected > 1:
            raise ApplicationError(
                f'يمكن اختيار وسم "{marker["label"]}" مرة واحدة فقط في فئة "{category["label"]}".'
            )
        if rule == "exactly_one_when_rows" and rows and selected != 1:
            raise ApplicationError(
                f'اختر بطاقة واحدة بوسم "{marker["label"]}" في فئة "{category["label"]}".'
            )
        if rule == "exactly_one_always" and selected != 1:
            raise ApplicationError(
                f'يجب اختيار بطاقة واحدة بوسم "{marker["label"]}" في فئة "{category["label"]}".'
            )




def _row_has_data(
    values: dict[str, Any],
    fields: Iterable[dict[str, Any]],
    markers: dict[str, Any] | None = None,
) -> bool:
    if any(bool(value) for value in (markers or {}).values()):
        return True
    for field in fields:
        value = values.get(field["id"], "")
        if field["type"] == "file" and isinstance(value, dict):
            if value.get("stored_path") or value.get("upload"):
                return True
            continue
        if field["type"] == "checkbox":
            if bool(value):
                return True
            continue
        if field["type"] == "checkbox_group":
            if parse_checkbox_group_value(value):
                return True
            continue
        if value is not None and clean_text(value):
            return True
    return False


def attachments_directory() -> Path:
    return DATA_DIR / "attachments"


def attachment_relative_path(value: Any) -> str:
    if isinstance(value, dict):
        value = value.get("stored_path", "")
    text = clean_text(value).replace("\\", "/")
    if not text:
        return ""
    if not text.startswith("attachments/"):
        raise ApplicationError("مسار ملف المرفق غير صالح.")
    filename = text.removeprefix("attachments/")
    if (
        not filename
        or filename in {".", ".."}
        or "/" in filename
        or "\\" in filename
    ):
        raise ApplicationError("مسار ملف المرفق غير صالح.")
    return f"attachments/{filename}"


def attachment_absolute_path(value: Any) -> Path | None:
    relative = attachment_relative_path(value)
    if not relative:
        return None
    return attachments_directory() / relative.removeprefix("attachments/")


def sanitize_filename_text(value: Any) -> str:
    text = unicodedata.normalize("NFKC", str(value or ""))
    text = INVALID_WINDOWS_FILENAME_PATTERN.sub(" ", text)
    return " ".join(text.split()).strip(" .")


def attachment_extension(original_name: Any) -> str:
    filename = str(original_name or "").replace("\\", "/").rsplit("/", 1)[-1]
    extension = Path(filename).suffix
    if (
        not extension
        or len(extension) > 20
        or INVALID_WINDOWS_FILENAME_PATTERN.search(extension)
        or any(character.isspace() for character in extension)
    ):
        return ""
    return extension


def decode_attachment_upload(upload: Any) -> tuple[str, bytes]:
    if not isinstance(upload, dict):
        raise ApplicationError("بيانات الملف غير صالحة.")
    original_name = clean_text(upload.get("name"))
    encoded = upload.get("data")
    if not original_name or not isinstance(encoded, str) or not encoded:
        raise ApplicationError("بيانات الملف غير مكتملة.")
    try:
        content = base64.b64decode(encoded, validate=True)
    except (ValueError, binascii.Error) as exc:
        raise ApplicationError("تعذّر قراءة الملف.") from exc
    if len(content) > MAX_ATTACHMENT_BYTES:
        raise ApplicationError("حجم الملف يتجاوز 100 ميغابايت.")
    return original_name, content


def _field_display_value(
    field_id: str,
    main_values: dict[str, Any],
    row_values: dict[str, Any] | None,
    indexes: dict[str, Any],
) -> str:
    if row_values and field_id in row_values:
        value = row_values.get(field_id, "")
    else:
        value = main_values.get(field_id, "")
    if isinstance(value, dict):
        return ""
    field = indexes["fields"].get(field_id)
    if field and field["type"] in {"select", "yes_no", "checkbox_group"}:
        options = {
            str(option["id"]): str(option["label"])
            for option in field.get("options", [])
        }
        if field["type"] == "checkbox_group":
            if isinstance(value, str):
                try:
                    parsed = json.loads(value)
                except (TypeError, ValueError, json.JSONDecodeError):
                    parsed = [value]
            else:
                parsed = value
            if isinstance(parsed, list):
                return "، ".join(
                    options.get(str(item), str(item)) for item in parsed
                )
        return options.get(str(value), clean_text(json_value(value)))
    return clean_text(json_value(value))


def attachment_base_name(
    field: dict[str, Any],
    original_name: str,
    main_values: dict[str, Any],
    row_values: dict[str, Any] | None,
    indexes: dict[str, Any],
) -> str:
    naming = field["file_naming"]
    if naming["mode"] == "original":
        base = sanitize_filename_text(Path(original_name).stem)
        if not base:
            raise ApplicationError("اسم الملف الأصلي غير صالح.")
        return base[:180]

    pieces: list[str] = []
    missing: list[str] = []
    for part in naming["parts"]:
        source_id = part["field_id"]
        value = _field_display_value(source_id, main_values, row_values, indexes)
        if not value:
            missing.append(indexes["fields"][source_id]["label"])
        prefix = str(part.get("prefix", ""))
        suffix = str(part.get("suffix", ""))

        pieces.append(
            f"{prefix}{value}{suffix}"
        )
    if missing:
        raise ApplicationError(
            f'لا يمكن رفع "{field["label"]}" قبل تعبئة: '
            + "، ".join(missing)
            + "."
        )
    base = sanitize_filename_text("".join(pieces))
    if not base:
        raise ApplicationError("تعذّر إنشاء اسم صالح للملف.")
    return base[:180].rstrip(" .")


def unique_attachment_destination(
    base: str,
    extension: str,
    reserved: set[Path],
) -> tuple[str, Path]:
    directory = attachments_directory()
    counter = 1
    while True:
        suffix = "" if counter == 1 else f" ({counter})"
        filename = f"{base}{suffix}{extension}"
        destination = directory / filename
        if destination not in reserved and not destination.exists():
            reserved.add(destination)
            return f"attachments/{filename}", destination
        counter += 1


def prepare_file_value(
    raw_value: Any,
    field: dict[str, Any],
    main_values: dict[str, Any],
    row_values: dict[str, Any] | None,
    indexes: dict[str, Any],
    mode: str,
    old_paths: set[str],
    kept_paths: set[str],
    staged: list[tuple[Path, Path]],
    reserved: set[Path],
) -> str:
    if not raw_value:
        return ""
    if not isinstance(raw_value, dict):
        raw_value = {"stored_path": raw_value}
    upload = raw_value.get("upload")
    stored_path = attachment_relative_path(raw_value.get("stored_path", ""))
    if upload:
        original_name, content = decode_attachment_upload(upload)
        base = attachment_base_name(
            field, original_name, main_values, row_values, indexes
        )
        extension = attachment_extension(original_name)
        relative, destination = unique_attachment_destination(
            base, extension, reserved
        )
        attachments_directory().mkdir(parents=True, exist_ok=True)
        with tempfile.NamedTemporaryFile(
            prefix=".attachment-",
            suffix=".tmp",
            dir=attachments_directory(),
            delete=False,
        ) as temporary:
            temporary.write(content)
            temporary_path = Path(temporary.name)
        staged.append((temporary_path, destination))
        kept_paths.add(relative)
        return relative
    if stored_path:
        if mode != "update" or stored_path not in old_paths:
            raise ApplicationError("مرجع الملف لا يعود إلى هذا السجل.")
        kept_paths.add(stored_path)
        return stored_path
    return ""


def remove_attachment_files(paths: set[str]) -> int:
    deleted = 0
    for value in paths:
        path = attachment_absolute_path(value)
        if path is None:
            continue
        try:
            path.unlink()
            deleted += 1
        except FileNotFoundError:
            continue
        except OSError:
            continue
    return deleted


def create_backup(*, automatic: bool = False) -> dict[str, str]:
    """Create a consistent ZIP of the schema, workbook, and attachments."""
    with WORKBOOK_LOCK:
        ensure_storage()
        BACKUP_DIR.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now().strftime("%Y%m%d-%H%M%S-%f")
        kind = "auto-backup" if automatic else "backup"
        filename = f"GenericSchemaCraft-{kind}-{stamp}.zip"
        destination = BACKUP_DIR / filename
        with tempfile.NamedTemporaryFile(
            prefix=".backup-",
            suffix=".zip",
            dir=BACKUP_DIR,
            delete=False,
        ) as temporary:
            temporary_path = Path(temporary.name)
        try:
            with zipfile.ZipFile(
                temporary_path,
                mode="w",
                compression=zipfile.ZIP_DEFLATED,
            ) as archive:
                archive.write(SCHEMA_PATH, "schema.json")
                archive.write(WORKBOOK_PATH, "database.xlsx")
                attachment_files = sorted(
                    path
                    for path in attachments_directory().rglob("*")
                    if path.is_file()
                )
                if not attachment_files:
                    archive.writestr("attachments/", b"")
                for path in attachment_files:
                    relative = path.relative_to(attachments_directory())
                    archive.write(
                        path,
                        (Path("attachments") / relative).as_posix(),
                    )
            os.replace(temporary_path, destination)
        finally:
            temporary_path.unlink(missing_ok=True)
        if automatic:
            automatic_backups = sorted(
                BACKUP_DIR.glob("GenericSchemaCraft-auto-backup-*.zip"),
                key=lambda path: path.stat().st_mtime,
                reverse=True,
            )
            for old in automatic_backups[10:]:
                old.unlink(missing_ok=True)
    return {
        "filename": filename,
        "download_url": f"/api/backups/{quote(filename)}",
        "automatic": automatic,
    }


def backup_file_path(filename: Any) -> Path | None:
    text = clean_text(filename)
    name = Path(text).name
    if name != text:
        return None
    if not re.fullmatch(
        r"GenericSchemaCraft-(?:auto-)?backup-\d{8}-\d{6}-\d{6}\.zip",
        name,
    ):
        return None
    return BACKUP_DIR / name


def required_value_missing(field: dict[str, Any], value: Any) -> bool:
    if field["type"] == "checkbox":
        return not bool(value)
    if field["type"] == "checkbox_group":
        return not bool(value)
    return value in {"", None}


def _normalize_submission(
    schema: dict[str, Any],
    payload: dict[str, Any],
    mode: str,
    old_paths: set[str],
) -> tuple[
    dict[str, Any],
    dict[str, list[dict[str, Any]]],
    list[tuple[Path, Path]],
    set[str],
]:
    indexes = schema_indexes(schema)
    raw_main = payload.get("main", {})
    raw_related = payload.get("related", {})
    if not isinstance(raw_main, dict) or not isinstance(raw_related, dict):
        raise ApplicationError("صيغة بيانات السجل غير صحيحة.")

    main_values = {
        field["id"]: raw_main.get(field["id"], "")
        for field in main_fields(schema)
    }
    staged: list[tuple[Path, Path]] = []
    kept_paths: set[str] = set()
    reserved: set[Path] = set()
    normalized_main: dict[str, Any] = {}
    normalized_related: dict[str, list[dict[str, Any]]] = {}
    visible_main_fields: list[dict[str, Any]] = []

    try:
        for category in schema["categories"]:
            if category["kind"] != "main":
                continue
            category_visible = target_visible(
                "category", category["id"], main_values, None, indexes
            )
            for field in category["fields"]:
                if field["type"] in SYSTEM_FIELD_TYPES:
                    continue
                visible = category_visible and target_visible(
                    "field",
                    field["id"],
                    main_values,
                    None,
                    indexes,
                )

                # Visibility must never erase an entered or previously saved value.
                raw = main_values.get(field["id"], "")

                if visible:
                    visible_main_fields.append(field)
                if field["type"] == "file":
                    normalized = prepare_file_value(
                        raw,
                        field,
                        main_values,
                        None,
                        indexes,
                        mode,
                        old_paths,
                        kept_paths,
                        staged,
                        reserved,
                    )
                else:
                    normalized = normalize_field_value(raw, field)

                    # A hidden dependent field may retain a value that is not
                    # currently allowed by its controlling field.
                    if visible:
                        validate_dependent_option_value(
                            field,
                            normalized,
                            main_values,
                            None,
                            indexes,
                        )
                if field["required"] and visible and required_value_missing(field, normalized):
                    raise ApplicationError(f'الحقل "{field["label"]}" مطلوب.')
                normalized_main[field["id"]] = normalized

        validate_cross_field_constraints(
            visible_main_fields,
            normalized_main,
        )

        for category in schema["categories"]:
            if category["kind"] != "repeatable":
                continue
            category_visible = target_visible(
                "category", category["id"], main_values, None, indexes
            )
            raw_rows = raw_related.get(category["id"], [])
            if not isinstance(raw_rows, list):
                rows: list[dict[str, Any]] = []
                validate_marker_rules(category, rows)
                normalized_related[category["id"]] = rows
                continue

            rows = []
            valid_marker_ids = {
                marker["id"] for marker in category.get("row_markers", [])
            }
            for raw_row in raw_rows:
                if not isinstance(raw_row, dict):
                    continue
                raw_values = raw_row.get("values", raw_row)
                raw_markers = raw_row.get("markers", {})
                if not isinstance(raw_values, dict):
                    continue
                if not isinstance(raw_markers, dict):
                    raw_markers = {}
                markers = {
                    marker_id: excel_boolean(raw_markers.get(marker_id, False))
                    for marker_id in valid_marker_ids
                }
                linked_record_code = clean_text(
                    raw_row.get("linked_record_code")
                )
                related_person_mode = clean_text(
                    raw_row.get("related_person_mode")
                )
                if linked_record_code:
                    related_person_mode = "existing"
                elif related_person_mode == "existing":
                    raise ApplicationError(
                        f'أدخل معرّف الشخص المرتبط في فئة "{category["label"]}".'
                    )
                else:
                    related_person_mode = "manual"
                if not _row_has_data(raw_values, category["fields"], markers) and not linked_record_code:
                    continue
                condition_row_values = dict(raw_values)
                mode_source = related_person_mode_source_field(category)
                if mode_source:
                    condition_row_values[mode_source["id"]] = related_person_mode
                values: dict[str, Any] = {}
                visible_row_fields: list[dict[str, Any]] = []
                for field in category["fields"]:
                    visible = category_visible and target_visible(
                        "field",
                        field["id"],
                        main_values,
                        condition_row_values,
                        indexes,
                    )

                    raw = raw_values.get(field["id"], "")

                    if visible:
                        visible_row_fields.append(field)
                    if field["type"] == "file":
                        normalized = prepare_file_value(
                            raw,
                            field,
                            main_values,
                            raw_values,
                            indexes,
                            mode,
                            old_paths,
                            kept_paths,
                            staged,
                            reserved,
                        )
                    else:
                        normalized = normalize_field_value(raw, field)

                        if visible:
                            validate_dependent_option_value(
                                field,
                                normalized,
                                main_values,
                                raw_values,
                                indexes,
                            )
                    if field["required"] and visible and required_value_missing(field, normalized):
                        raise ApplicationError(
                            f'الحقل "{field["label"]}" مطلوب في فئة '
                            f'"{category["label"]}".'
                        )
                    values[field["id"]] = normalized
                validate_cross_field_constraints(
                    visible_row_fields,
                    values,
                )
                child_id = clean_text(raw_row.get("_child_id"))
                if child_id:
                    child_id = validate_internal_id(child_id, "معرّف الصف")
                rows.append(
                    {
                        "_child_id": child_id,
                        "linked_record_code": linked_record_code,
                        "values": values,
                        "markers": markers,
                    }
                )
            validate_marker_rules(category, rows)
            normalized_related[category["id"]] = rows
    except Exception:
        for temporary_path, _ in staged:
            temporary_path.unlink(missing_ok=True)
        raise

    return normalized_main, normalized_related, staged, kept_paths


def validate_related_person_links(
    schema: dict[str, Any],
    records: list[dict[str, Any]],
    current_record_code: str,
    related_values: dict[str, list[dict[str, Any]]],
) -> None:
    known_codes = {record["record_code"] for record in records}
    for category in schema["categories"]:
        if category["kind"] != "repeatable":
            continue
        enabled = bool(category.get("related_person_enabled"))
        for row in related_values.get(category["id"], []):
            linked_code = clean_text(row.get("linked_record_code"))
            if not linked_code:
                continue
            if not enabled:
                raise ApplicationError(
                    f'الفئة "{category["label"]}" غير مهيأة لربط الأشخاص.'
                )
            linked_code = validate_person_code(linked_code)
            if linked_code == current_record_code:
                raise ApplicationError("لا يمكن ربط السجل بنفسه كشخص مرتبط.")
            if linked_code not in known_codes:
                raise ApplicationError(
                    f'لم يُعثر على الشخص المرتبط ذي المعرّف "{linked_code}".'
                )
            row["linked_record_code"] = linked_code


def _find_record(
    records: list[dict[str, Any]], record_code: str
) -> dict[str, Any] | None:
    return next(
        (record for record in records if record["record_code"] == record_code),
        None,
    )


def canonical_unique_value(value: Any) -> str:
    if isinstance(value, list):
        return json.dumps(
            sorted(normalize_search_text(item) for item in value),
            ensure_ascii=False,
        )
    return normalize_search_text(value)


def unique_field_values_in_record(
    record: dict[str, Any], category: dict[str, Any], field: dict[str, Any]
) -> list[Any]:
    if category["kind"] == "main":
        return [record.get("values", {}).get(field["id"], "")]
    return [
        row.get("values", {}).get(field["id"], "")
        for row in record.get("related", {}).get(category["id"], [])
    ]


def validate_unique_fields(
    schema: dict[str, Any],
    records: list[dict[str, Any]],
    current_record_id: str | None,
    main_values: dict[str, Any],
    related_values: dict[str, list[dict[str, Any]]],
    unique_indexes: dict[str, dict[str, frozenset[str]]] | None = None,
) -> None:
    for category in schema["categories"]:
        for field in category["fields"]:
            if not field.get("unique"):
                continue
            if category["kind"] == "main":
                submitted = [main_values.get(field["id"], "")]
            else:
                submitted = [
                    row.get("values", {}).get(field["id"], "")
                    for row in related_values.get(category["id"], [])
                ]
            submitted_tokens = [
                canonical_unique_value(value)
                for value in submitted
                if value != "" and value is not None and value != []
            ]
            if len(submitted_tokens) != len(set(submitted_tokens)):
                raise ApplicationError(
                    f'قيمة الحقل الفريد "{field["label"]}" مكررة داخل السجل.'
                )
            if not submitted_tokens:
                continue

            if unique_indexes is not None:
                field_index = unique_indexes.get(field["id"], {})
                for token in submitted_tokens:
                    owners = field_index.get(token, frozenset())
                    if any(owner != current_record_id for owner in owners):
                        raise ApplicationError(
                            f'قيمة الحقل "{field["label"]}" مستخدمة في سجل آخر.'
                        )
                continue

            existing_tokens: set[str] = set()
            for record in records:
                if current_record_id and record.get("_record_id") == current_record_id:
                    continue
                for value in unique_field_values_in_record(record, category, field):
                    if value == "" or value is None or value == []:
                        continue
                    existing_tokens.add(canonical_unique_value(value))
            if any(token in existing_tokens for token in submitted_tokens):
                raise ApplicationError(
                    f'قيمة الحقل "{field["label"]}" مستخدمة في سجل آخر.'
                )




def save_record(payload: Any) -> dict[str, Any]:
    if not isinstance(payload, dict):
        raise ApplicationError("صيغة البيانات المرسلة غير صحيحة.")
    mode = payload.get("mode", "create")
    if mode not in {"create", "update"}:
        raise ApplicationError("نوع عملية الحفظ غير صالح.")
    requested_code = clean_text(payload.get("record_code"))
    record_code = validate_person_code(requested_code) if requested_code else ""

    operation_started = time.perf_counter()
    staged: list[tuple[Path, Path]] = []
    promoted: list[Path] = []
    obsolete_paths: set[str] = set()
    with WORKBOOK_LOCK:
        schema = read_schema_file()
        snapshot = _dataset_snapshot_unlocked(schema)
        records = list(snapshot.records)
        existing = snapshot.records_by_code.get(record_code) if record_code else None

        if mode == "create":
            if record_code and record_code in snapshot.records_by_code:
                raise ApplicationError(
                    "معرّف السجل مستخدم مسبقًا. افتح نموذجًا جديدًا للحصول على معرّف آخر."
                )
            if not record_code:
                record_code = generate_person_code()
                while record_code in snapshot.records_by_code:
                    record_code = generate_person_code()
            existing = None
        elif not record_code:
            raise ApplicationError("اختر سجلًا قبل حفظ التعديلات.")
        elif existing is None:
            raise ApplicationError("لم يُعثر على السجل المطلوب تعديله.")

        old_paths = _record_file_paths(schema, existing) if existing else set()
        try:
            normalize_started = time.perf_counter()
            (
                main_values,
                related_values,
                staged,
                kept_paths,
            ) = _normalize_submission(schema, payload, mode, old_paths)
            normalize_seconds = time.perf_counter() - normalize_started

            validate_started = time.perf_counter()
            validate_unique_fields(
                schema,
                records,
                existing.get("_record_id") if existing else None,
                main_values,
                related_values,
                snapshot.unique_indexes,
            )
            validate_related_person_links(
                schema,
                records,
                record_code,
                related_values,
            )
            validate_seconds = time.perf_counter() - validate_started
            timestamp = now_iso()

            if existing is None:
                record = {
                    "_record_id": new_internal_id(),
                    "record_code": record_code,
                    "created_at": timestamp,
                    "updated_at": timestamp,
                    "archived": False,
                    "archived_at": "",
                    "values": main_values,
                    "related": {},
                }
                records.append(record)
            else:
                # Never mutate an object owned by the published cache before
                # the Excel replacement succeeds.
                record = copy.deepcopy(existing)
                record["updated_at"] = timestamp
                record["values"] = main_values
                records[snapshot.record_positions[record["_record_id"]]] = record

            for category in schema["categories"]:
                if category["kind"] != "repeatable":
                    continue
                previous_rows = {
                    child["_child_id"]: child
                    for child in record.get("related", {}).get(category["id"], [])
                }
                new_rows = []
                for row in related_values.get(category["id"], []):
                    child_id = row["_child_id"]
                    if child_id and child_id not in previous_rows:
                        raise ApplicationError(
                            f'معرّف صف في الفئة "{category["label"]}" لا يعود إلى هذا السجل.'
                        )
                    previous = previous_rows.get(child_id)
                    new_rows.append(
                        {
                            "_child_id": child_id or new_internal_id(),
                            "minor_id": len(new_rows) + 1,
                            "created_at": (
                                previous["created_at"] if previous else timestamp
                            ),
                            "updated_at": timestamp,
                            "linked_record_code": row.get(
                                "linked_record_code", ""
                            ),
                            "markers": row.get("markers", {}),
                            "values": row["values"],
                        }
                    )
                record.setdefault("related", {})[category["id"]] = new_rows

            attachment_started = time.perf_counter()
            for temporary_path, destination in staged:
                os.replace(temporary_path, destination)
                promoted.append(destination)
            attachment_seconds = time.perf_counter() - attachment_started

            workbook_started = time.perf_counter()
            atomic_write_workbook(schema, records)
            workbook_seconds = time.perf_counter() - workbook_started
            obsolete_paths = old_paths - kept_paths
            deleted_files = remove_attachment_files(obsolete_paths)
            _safe_publish_dataset_snapshot(schema, records)
            LOGGER.info(
                "Record save timings: mode=%s normalize=%.3f validate=%.3f "
                "attachments=%.3f workbook=%.3f total=%.3f",
                mode,
                normalize_seconds,
                validate_seconds,
                attachment_seconds,
                workbook_seconds,
                time.perf_counter() - operation_started,
            )
        except PermissionError as exc:
            for temporary_path, _ in staged:
                temporary_path.unlink(missing_ok=True)
            for destination in promoted:
                destination.unlink(missing_ok=True)
            invalidate_dataset_cache()
            raise ApplicationError(
                "تعذّر الحفظ. أغلق database.xlsx في Excel ثم حاول مرة أخرى."
            ) from exc
        except Exception:
            for temporary_path, _ in staged:
                temporary_path.unlink(missing_ok=True)
            for destination in promoted:
                destination.unlink(missing_ok=True)
            invalidate_dataset_cache()
            raise

    attachment_paths = sorted(_record_file_paths(schema, record))
    related_count = sum(
        len(record.get("related", {}).get(category["id"], []))
        for category in schema["categories"]
        if category["kind"] == "repeatable"
    )
    return {
        "ok": True,
        "action": "updated" if mode == "update" else "created",
        "record_code": record_code,
        "_record_id": record["_record_id"],
        "related_rows": related_count,
        "attachment_files": attachment_paths,
        "deleted_attachment_files": deleted_files,
    }


def load_record(record_code: Any) -> dict[str, Any]:
    code = validate_person_code(record_code)
    schema = read_schema_file()
    with WORKBOOK_LOCK:
        record = _dataset_snapshot_unlocked(schema).records_by_code.get(code)
    if record is None:
        raise ApplicationError("لم يُعثر على السجل المطلوب.")
    return {
        "record_code": code,
        "_record_id": record["_record_id"],
        "created_at": record.get("created_at", ""),
        "updated_at": record.get("updated_at", ""),
        "archived": bool(record.get("archived")),
        "main": copy.deepcopy(record["values"]),
        "related": {
            category["id"]: [
                {
                    "_child_id": child["_child_id"],
                    "minor_id": index + 1,
                    "linked_record_code": child.get(
                        "linked_record_code", ""
                    ),
                    "markers": copy.deepcopy(child.get("markers", {})),
                    "values": copy.deepcopy(child["values"]),
                }
                for index, child in enumerate(
                    record.get("related", {}).get(category["id"], [])
                )
            ]
            for category in schema["categories"]
            if category["kind"] == "repeatable"
        },
    }


def _field_matches(field: dict[str, Any], query: Any, value: Any) -> bool:
    if field["type"] == "checkbox":
        return normalize_field_value(query, field) == normalize_field_value(value, field)
    if field["type"] == "checkbox_group":
        try:
            expected = {
                normalize_search_text(item)
                for item in normalize_field_value(query, field)
            }
        except ApplicationError:
            return False
        actual = {
            normalize_search_text(item)
            for item in parse_checkbox_group_value(value)
        }
        return expected.issubset(actual)
    if field["type"] in {"select", "yes_no"}:
        try:
            query = normalize_choice_value(query, field)
        except ApplicationError:
            return False
    if field["search_match"] == "contains":
        return normalize_search_text(query) in normalize_search_text(value)
    if field["type"] == "number":
        try:
            return float(query) == float(value)
        except (TypeError, ValueError):
            return False
    return normalize_search_text(query) == normalize_search_text(value)


def _indexed_field_matches(
    field: dict[str, Any], query: Any, indexed_value: Any
) -> bool:
    if field["type"] == "checkbox":
        return normalize_field_value(query, field) == indexed_value
    if field["type"] == "checkbox_group":
        try:
            expected = {
                normalize_search_text(item)
                for item in normalize_field_value(query, field)
            }
        except ApplicationError:
            return False
        return expected.issubset(indexed_value or frozenset())
    if field["type"] in {"select", "yes_no"}:
        try:
            query = normalize_choice_value(query, field)
        except ApplicationError:
            return False
    if field["search_match"] == "contains":
        return normalize_search_text(query) in (indexed_value or "")
    if field["type"] == "number":
        try:
            return float(query) == indexed_value
        except (TypeError, ValueError):
            return False
    return normalize_search_text(query) == (indexed_value or "")


def field_display_value(field: dict[str, Any], value: Any) -> Any:
    if field["type"] == "checkbox":
        return "نعم" if excel_boolean(value) else "لا"
    option_labels = {
        option["id"]: option["label"] for option in field.get("options", [])
    }
    if field["type"] == "checkbox_group":
        return "، ".join(
            option_labels.get(clean_text(item), clean_text(item))
            for item in parse_checkbox_group_value(value)
        )
    if field["type"] in {"select", "yes_no"}:
        return option_labels.get(clean_text(value), clean_text(value))
    return json_value(value)


def _field_values_for_record(
    record: dict[str, Any],
    category: dict[str, Any],
    field: dict[str, Any],
) -> list[Any]:
    if category["kind"] == "main":
        return [record.get("values", {}).get(field["id"], "")]
    return [
        row.get("values", {}).get(field["id"], "")
        for row in record.get("related", {}).get(category["id"], [])
    ]


def search_records(criteria: Any) -> dict[str, Any]:
    if not isinstance(criteria, dict):
        raise ApplicationError("صيغة البحث غير صحيحة.")
    schema = read_schema_file()
    indexes = schema_indexes(schema)
    requested_field_ids = criteria.get("_search_field_ids")
    if requested_field_ids is None:
        selected_field_ids = [
            field_id
            for field_id, field in indexes["fields"].items()
            if field["searchable"]
        ]
    else:
        if not isinstance(requested_field_ids, list):
            raise ApplicationError("حقول البحث المؤقتة غير صحيحة.")
        selected_field_ids = []
        for raw_field_id in requested_field_ids:
            field_id = clean_text(raw_field_id)
            field = indexes["fields"].get(field_id)
            if not field or field["type"] == "file":
                raise ApplicationError("أحد حقول البحث المؤقتة غير متاح.")
            if field_id not in selected_field_ids:
                selected_field_ids.append(field_id)
    searchable = {
        field_id: indexes["fields"][field_id]
        for field_id in selected_field_ids
    }
    include_archived = bool(criteria.get("_include_archived"))
    active = {
        field_id: value
        for field_id, value in criteria.items()
        if field_id in searchable
        and (
            bool(value)
            if searchable[field_id]["type"] in {"checkbox", "checkbox_group"}
            else bool(clean_text(value))
        )
    }
    if not active:
        raise ApplicationError("أدخل معيار بحث واحدًا على الأقل.")

    main_criteria: list[tuple[dict[str, Any], Any]] = []
    related_criteria: dict[str, list[tuple[dict[str, Any], Any]]] = {}
    for field_id, query in active.items():
        category_id = indexes["field_categories"][field_id]
        category = indexes["categories"][category_id]
        pair = (searchable[field_id], query)
        if category["kind"] == "main":
            main_criteria.append(pair)
        else:
            related_criteria.setdefault(category_id, []).append(pair)

    result_fields = [
        field
        for category in schema["categories"]
        for field in category["fields"]
        if field["show_in_results"]
    ]
    if not result_fields:
        result_fields = [
            field
            for category in schema["categories"]
            if category["kind"] == "main"
            for field in category["fields"]
            if field["type"] != "file"
        ][:4]

    matches = []
    truncated = False
    search_started = time.perf_counter()
    with WORKBOOK_LOCK:
        snapshot = _dataset_snapshot_unlocked(schema)
        records = snapshot.records

    for record in records:
        if record.get("archived") and not include_archived:
            continue
        record_id = record["_record_id"]
        main_index = snapshot.search_main.get(record_id, {})
        if any(
            not _indexed_field_matches(
                field, query, main_index.get(field["id"])
            )
            for field, query in main_criteria
        ):
            continue

        related_match = True
        related_index = snapshot.search_related.get(record_id, {})
        for category_id, pairs in related_criteria.items():
            indexed_rows = related_index.get(category_id, ())
            if not any(
                all(
                    _indexed_field_matches(
                        field, query, row.get(field["id"])
                    )
                    for field, query in pairs
                )
                for row in indexed_rows
            ):
                related_match = False
                break
        if not related_match:
            continue

        if len(matches) >= MAX_SEARCH_RESULTS:
            truncated = True
            break

        details = []
        title_parts = []
        title_field_ids: list[str] = []
        for field in result_fields:
            category = indexes["categories"][
                indexes["field_categories"][field["id"]]
            ]
            values = _field_values_for_record(record, category, field)
            value = next(
                (
                    field_display_value(field, item)
                    for item in values
                    if item != "" and item is not None and item != []
                ),
                "",
            )
            if field["result_title"] and value:
                title_parts.append(str(value))
                title_field_ids.append(field["id"])
            details.append(
                {
                    "field_id": field["id"],
                    "label": field["label"],
                    "value": value,
                }
            )
        if not title_parts:
            fallback_title_items = [
                item for item in details if clean_text(item["value"])
            ][:2]
            title_parts = [str(item["value"]) for item in fallback_title_items]
            title_field_ids = [item["field_id"] for item in fallback_title_items]
        detail_items = [
            item for item in details if item["field_id"] not in title_field_ids
        ]
        matches.append(
            {
                "record_code": record["record_code"],
                "title": " ".join(title_parts) or record["record_code"],
                "archived": bool(record.get("archived")),
                "details": detail_items,
            }
        )
    LOGGER.info(
        "Search completed: %.3f seconds; scanned=%d; matches=%d; criteria=%d",
        time.perf_counter() - search_started,
        len(records),
        len(matches),
        len(active),
    )
    return {"matches": matches, "truncated": truncated}


def archive_record(record_code: Any, archived: bool = True) -> dict[str, Any]:
    code = validate_person_code(record_code)
    with WORKBOOK_LOCK:
        schema = read_schema_file()
        snapshot = _dataset_snapshot_unlocked(schema)
        cached_record = snapshot.records_by_code.get(code)
        if cached_record is None:
            raise ApplicationError("لم يُعثر على السجل المطلوب.")
        records = list(snapshot.records)
        record = copy.deepcopy(cached_record)
        records[snapshot.record_positions[record["_record_id"]]] = record
        record["archived"] = bool(archived)
        record["archived_at"] = now_iso() if archived else ""
        record["updated_at"] = now_iso()
        try:
            atomic_write_workbook(schema, records)
            _safe_publish_dataset_snapshot(schema, records)
        except PermissionError as exc:
            invalidate_dataset_cache()
            raise ApplicationError(
                "تعذّر تحديث حالة الأرشفة. أغلق database.xlsx في Excel ثم حاول مرة أخرى."
            ) from exc
    return {
        "ok": True,
        "record_code": code,
        "archived": bool(archived),
        "archived_at": record["archived_at"],
    }




def delete_record(record_code: Any) -> dict[str, Any]:
    require_builder_access()
    code = validate_person_code(record_code)
    with WORKBOOK_LOCK:
        schema = read_schema_file()
        snapshot = _dataset_snapshot_unlocked(schema)
        record = snapshot.records_by_code.get(code)
        if record is None:
            raise ApplicationError("لم يُعثر على السجل المطلوب حذفه.")
        paths = _record_file_paths(schema, record)
        related_rows = sum(
            len(record.get("related", {}).get(category["id"], []))
            for category in schema["categories"]
            if category["kind"] == "repeatable"
        )
        records = [
            candidate
            for candidate in snapshot.records
            if candidate["_record_id"] != record["_record_id"]
        ]
        try:
            atomic_write_workbook(schema, records)
            _safe_publish_dataset_snapshot(schema, records)
        except PermissionError as exc:
            invalidate_dataset_cache()
            raise ApplicationError(
                "تعذّر الحذف. أغلق database.xlsx في Excel ثم حاول مرة أخرى."
            ) from exc
        deleted_files = remove_attachment_files(paths)
    return {
        "ok": True,
        "record_code": code,
        "deleted_main_rows": 1,
        "deleted_related_rows": related_rows,
        "deleted_attachment_files": deleted_files,
    }



class DataEntryHTTPServer(ThreadingHTTPServer):
    daemon_threads = True
    allow_reuse_address = True

    def __init__(self, server_address, handler_class):
        super().__init__(server_address, handler_class)

        self._browser_session_lock = threading.RLock()
        self._last_browser_heartbeat = 0.0

    def note_browser_heartbeat(self) -> None:
        with self._browser_session_lock:
            self._last_browser_heartbeat = time.monotonic()

    def browser_is_active(self) -> bool:
        with self._browser_session_lock:
            last_seen = self._last_browser_heartbeat

        if last_seen <= 0:
            return False

        return (
            time.monotonic() - last_seen
            <= BROWSER_ACTIVE_SECONDS
        )

    def touch_client(self) -> None:
        # Kept for compatibility with existing request code.
        # Normal API requests do not control application shutdown.
        return

    def note_disconnect(self) -> None:
        # Closing a browser page marks it inactive, but does not stop
        # the Python server.
        with self._browser_session_lock:
            self._last_browser_heartbeat = 0.0

    def handle_error(self, request, client_address) -> None:
        LOGGER.exception(
            "Unhandled request error from %s:%s",
            client_address[0],
            client_address[1],
        )

class DataEntryRequestHandler(SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=str(APP_DIR), **kwargs)

    def log_message(self, _format: str, *args) -> None:
        return

    def guess_type(self, path: str) -> str:
        content_type = super().guess_type(path)
        suffix = Path(urlparse(path).path).suffix.casefold()

        if suffix == ".html":
            return "text/html; charset=utf-8"
        if suffix == ".css":
            return "text/css; charset=utf-8"
        if suffix == ".js":
            return "text/javascript; charset=utf-8"

        return content_type

    def request_origin_is_allowed(self) -> bool:
        origin = self.headers.get("Origin")
        if not origin:
            return True

        expected_port = self.server.server_address[1]
        allowed_origins = {
            f"http://127.0.0.1:{expected_port}",
            f"http://localhost:{expected_port}",
        }
        return origin.rstrip("/") in allowed_origins

    def reject_foreign_origin(self) -> bool:
        if self.request_origin_is_allowed():
            return False

        LOGGER.warning(
            "Rejected foreign request origin: %s",
            self.headers.get("Origin"),
        )
        self.send_json(403, {"error": "مصدر الطلب غير مسموح."})
        return True

    def end_headers(self) -> None:
        path = unquote(urlparse(self.path).path)

        if not path.startswith("/api/"):
            self.send_header(
                "Cache-Control",
                "no-store, no-cache, must-revalidate, max-age=0",
            )
            self.send_header("Pragma", "no-cache")
            self.send_header("Expires", "0")

        super().end_headers()

    def send_json(self, status: int, body: dict[str, Any]) -> None:
        payload = json.dumps(body, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(payload)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(payload)

    def touch_client(self) -> None:
        if isinstance(self.server, DataEntryHTTPServer):
            self.server.touch_client()

    def _read_json(self) -> Any:
        try:
            length = int(self.headers.get("Content-Length", "0"))
        except ValueError as exc:
            raise ApplicationError("حجم الطلب غير صالح.") from exc
        if length > MAX_REQUEST_BYTES:
            raise ApplicationError("حجم الطلب يتجاوز الحد المسموح.")
        raw = self.rfile.read(length)
        try:
            return json.loads(raw.decode("utf-8"))
        except (UnicodeDecodeError, json.JSONDecodeError) as exc:
            raise ApplicationError("صيغة JSON غير صحيحة.") from exc

    def do_GET(self) -> None:
        path = unquote(urlparse(self.path).path)
        if path == "/api/health":
            browser_active = (
                self.server.browser_is_active()
                if isinstance(self.server, DataEntryHTTPServer)
                else False
            )

            self.send_json(
                200,
                {
                    "ok": True,
                    "browser_active": browser_active,
                },
            )
            return
        if path == "/api/schema":
            self.touch_client()
            try:
                self.send_json(
                    200,
                    schema_response(read_schema_with_excel_labels()),
                )
            except ApplicationError as exc:
                self.send_json(400, {"error": str(exc)})
            return
        if path.startswith("/api/records/"):
            self.touch_client()
            try:
                code = path.removeprefix("/api/records/")
                self.send_json(200, load_record(code))
            except ApplicationError as exc:
                self.send_json(400, {"error": str(exc)})
            return
        if path.startswith("/api/attachments/"):
            self.touch_client()
            try:
                filename = path.removeprefix("/api/attachments/")
                attachment = attachment_absolute_path(f"attachments/{filename}")
                if attachment is None or not attachment.is_file():
                    self.send_error(404)
                    return
                content = attachment.read_bytes()
                inline = (
                    attachment.suffix.casefold()
                    in INLINE_ATTACHMENT_EXTENSIONS
                )
                content_type = (
                    mimetypes.guess_type(attachment.name)[0]
                    if inline
                    else "application/octet-stream"
                ) or "application/octet-stream"
                self.send_response(200)
                self.send_header("Content-Type", content_type)
                self.send_header("Content-Length", str(len(content)))
                self.send_header(
                    "Content-Disposition",
                    f"{'inline' if inline else 'attachment'}; "
                    f"filename*=UTF-8''{quote(attachment.name)}",
                )
                self.send_header("X-Content-Type-Options", "nosniff")
                if not inline:
                    self.send_header("Content-Security-Policy", "sandbox")
                    self.send_header("X-Download-Options", "noopen")
                self.end_headers()
                self.wfile.write(content)
            except (ApplicationError, OSError):
                self.send_error(404)
            return
        if path.startswith("/api/backups/"):
            self.touch_client()
            if not builder_is_unlocked():
                self.send_error(404)
                return
            backup = backup_file_path(path.removeprefix("/api/backups/"))
            if backup is None or not backup.is_file():
                self.send_error(404)
                return
            try:
                content = backup.read_bytes()
                self.send_response(200)
                self.send_header("Content-Type", "application/zip")
                self.send_header("Content-Length", str(len(content)))
                self.send_header(
                    "Content-Disposition",
                    f"attachment; filename*=UTF-8''{quote(backup.name)}",
                )
                self.send_header("X-Content-Type-Options", "nosniff")
                self.end_headers()
                self.wfile.write(content)
            except OSError:
                self.send_error(404)
            return
        super().do_GET()

    def do_POST(self) -> None:
        if self.reject_foreign_origin():
            return

        path = unquote(urlparse(self.path).path)
        if path == "/api/heartbeat":
            if isinstance(self.server, DataEntryHTTPServer):
                self.server.note_browser_heartbeat()

            self.send_json(
                200,
                {
                    "ok": True,
                    "browser_active": True,
                },
            )
            return
        if path == "/api/disconnect":
            if isinstance(self.server, DataEntryHTTPServer):
                self.server.note_disconnect()

            try:
                self.send_json(200, {"ok": True})
            except (
                BrokenPipeError,
                ConnectionResetError,
            ):
                pass

            return
        if path == "/api/shutdown":
            self.send_json(200, {"ok": True})
            threading.Thread(target=self.server.shutdown, daemon=True).start()
            return
        if path == "/api/backup":
            self.touch_client()
            try:
                require_builder_access()
                self.send_json(200, create_backup())
            except ApplicationError as exc:
                LOGGER.warning(
                    "Backup request rejected: %s",
                    exc,
                )

                self.send_json(
                    400,
                    {"error": str(exc)},
                )

            except Exception:
                LOGGER.exception(
                    "Unexpected backup creation failure"
                )

                self.send_json(
                    500,
                    {
                        "error":
                            "تعذّر إنشاء النسخة الاحتياطية بسبب خطأ داخلي."
                    },
                )
            return
        try:
            self.touch_client()
            payload = self._read_json()
            if path == "/api/client-log":
                level = clean_text(
                    payload.get("level", "error")
                ).casefold()

                category = clean_text(
                    payload.get("category", "browser")
                )[:80]

                message = clean_text(
                    payload.get("message", "Unknown browser error")
                )[:4000]

                location = clean_text(
                    payload.get("location", "")
                )[:4000]

                occurred_at = clean_text(
                    payload.get("occurred_at", "")
                )[:80]

                log_text = (
                    f"Browser [{category}] {message}"
                )

                if occurred_at:
                    log_text += f" | occurred_at={occurred_at}"

                if location:
                    log_text += f"\n{location}"

                if level == "warning":
                    LOGGER.warning(log_text)
                else:
                    LOGGER.error(log_text)

                self.send_json(
                    200,
                    {"ok": True},
                )
                return
            if path == "/api/builder/unlock":
                access = unlock_builder(
                    payload.get("password"),
                    initialize=bool(payload.get("initialize")),
                )
                self.send_json(200, {"ok": True, "builder_access": access})
                return
            if path == "/api/builder/lock":
                self.send_json(200, {"ok": True, "builder_access": lock_builder()})
                return
            if path == "/api/builder/password":
                self.send_json(
                    200,
                    {
                        "ok": True,
                        "builder_access": change_builder_password(
                            payload.get("current_password"),
                            payload.get("new_password"),
                        ),
                    },
                )
                return
            if path == "/api/records":
                self.send_json(200, save_record(payload))
                return
            if path == "/api/archive":
                self.send_json(
                    200,
                    archive_record(
                        payload.get("record_code"),
                        bool(payload.get("archived", True)),
                    ),
                )
                return
            if path == "/api/search":
                self.send_json(200, search_records(payload))
                return
            self.send_json(404, {"error": "المسار المطلوب غير موجود."})
        except ApplicationError as exc:
            LOGGER.warning(
                "Request rejected: %s %s | %s",
                self.command,
                path,
                exc,
            )

            self.send_json(
                400,
                {"error": str(exc)},
            )

        except Exception:
            LOGGER.exception(
                "Unexpected API error: %s %s",
                self.command,
                path,
            )

            self.send_json(
                500,
                {"error": "حدث خطأ داخلي غير متوقع في التطبيق."},
            )

    def do_PUT(self) -> None:
        if self.reject_foreign_origin():
            return

        path = unquote(urlparse(self.path).path)
        try:
            self.touch_client()
            if path != "/api/schema":
                self.send_json(404, {"error": "المسار المطلوب غير موجود."})
                return
            require_builder_access()
            self.send_json(200, save_schema(self._read_json()))
        except ApplicationError as exc:
            LOGGER.warning(
                "Request rejected: %s %s | %s",
                self.command,
                path,
                exc,
            )
            self.send_json(400, {"error": str(exc)})
        except Exception:
            LOGGER.exception(
                "Unexpected API error: %s %s",
                self.command,
                path,
            )
            self.send_json(
                500,
                {"error": "حدث خطأ داخلي غير متوقع في التطبيق."},
            )

    def do_DELETE(self) -> None:
        if self.reject_foreign_origin():
            return

        path = unquote(urlparse(self.path).path)
        self.touch_client()
        try:
            if not path.startswith("/api/records/"):
                self.send_json(404, {"error": "المسار المطلوب غير موجود."})
                return
            code = path.removeprefix("/api/records/")
            self.send_json(200, delete_record(code))
        except ApplicationError as exc:
            LOGGER.warning(
                "Request rejected: %s %s | %s",
                self.command,
                path,
                exc,
            )

            self.send_json(
                400,
                {"error": str(exc)},
            )

        except Exception:
            LOGGER.exception(
                "Unexpected API error: %s %s",
                self.command,
                path,
            )

            self.send_json(
                500,
                {"error": "حدث خطأ داخلي غير متوقع في التطبيق."},
            )


def acquire_single_instance() -> bool:
    """Return False when this application folder is already running."""
    global _WINDOWS_MUTEX_HANDLE

    if os.name != "nt":
        return True

    folder_key = hashlib.sha256(
        str(BASE_DIR).casefold().encode("utf-8")
    ).hexdigest()[:20]

    mutex_name = f"Local\\GenericSchemaCraft-{folder_key}"

    kernel32 = ctypes.windll.kernel32

    kernel32.CreateMutexW.argtypes = [
        ctypes.c_void_p,
        ctypes.c_bool,
        ctypes.c_wchar_p,
    ]
    kernel32.CreateMutexW.restype = ctypes.c_void_p

    kernel32.CloseHandle.argtypes = [ctypes.c_void_p]
    kernel32.CloseHandle.restype = ctypes.c_bool

    handle = kernel32.CreateMutexW(None, False, mutex_name)

    if not handle:
        raise OSError("تعذّر إنشاء قفل تشغيل البرنامج.")

    if kernel32.GetLastError() == WINDOWS_ALREADY_EXISTS:
        kernel32.CloseHandle(ctypes.c_void_p(handle))
        return False

    _WINDOWS_MUTEX_HANDLE = int(handle)
    return True

def open_browser_when_needed(
    server: DataEntryHTTPServer,
    url: str,
) -> None:
    """
    Wait for an old browser page to reconnect.

    Open a new page only when no browser session appears.
    """
    deadline = (
        time.monotonic()
        + STARTUP_BROWSER_WAIT_SECONDS
    )

    while time.monotonic() < deadline:
        if server.browser_is_active():
            LOGGER.info(
                "Existing browser session reconnected; "
                "no new page was opened."
            )
            return

        time.sleep(0.25)

    if server.browser_is_active():
        return

    LOGGER.info(
        "No active browser session detected; opening %s",
        url,
    )

    open_application_window(url)


def _application_browser_candidates() -> list[Path]:
    """Return installed Chromium-family browsers that support --app mode."""
    candidates: list[Path] = []
    commands = (
        "msedge",
        "microsoft-edge",
        "google-chrome",
        "chrome",
        "chromium",
        "chromium-browser",
    )
    for command in commands:
        executable = shutil.which(command)
        if executable:
            candidates.append(Path(executable))

    if os.name == "nt":
        roots = [
            os.environ.get("PROGRAMFILES(X86)"),
            os.environ.get("PROGRAMFILES"),
            os.environ.get("LOCALAPPDATA"),
        ]
        relative_paths = (
            Path("Microsoft/Edge/Application/msedge.exe"),
            Path("Google/Chrome/Application/chrome.exe"),
        )
        for root in roots:
            if not root:
                continue
            for relative_path in relative_paths:
                candidate = Path(root) / relative_path
                if candidate.is_file():
                    candidates.append(candidate)

    unique: list[Path] = []
    seen: set[str] = set()
    for candidate in candidates:
        key = str(candidate).casefold()
        if key not in seen:
            seen.add(key)
            unique.append(candidate)
    return unique


def open_application_window(url: str) -> None:
    """Prefer a clean app window, while retaining the normal-browser fallback."""
    for executable in _application_browser_candidates():
        try:
            subprocess.Popen(
                [str(executable), f"--app={url}", "--new-window"],
                close_fds=True,
            )
            LOGGER.info("Opened application-mode browser window with %s", executable)
            return
        except OSError:
            LOGGER.exception("Could not open application-mode browser %s", executable)

    LOGGER.info("No application-mode browser found; using the default browser.")
    webbrowser.open(url)
def open_running_instance() -> bool:
    """
    Connect to the already-running local server.

    Open a browser page only if the server has no active page.
    """
    port = application_port()
    url = f"http://{HOST}:{port}/"
    health_url = f"{url}api/health"

    health: dict[str, Any] | None = None

    # The first instance may still be starting.
    for _attempt in range(20):
        try:
            with urlopen(
                health_url,
                timeout=0.5,
            ) as response:
                if response.status != 200:
                    time.sleep(0.15)
                    continue

                raw = response.read().decode("utf-8")
                health = json.loads(raw or "{}")
                break

        except (
            OSError,
            URLError,
            json.JSONDecodeError,
        ):
            time.sleep(0.15)

    if health is None:
        return False

    if health.get("browser_active", False):
        LOGGER.info(
            "Application and browser session are already open."
        )
        return True

    # Give a recently opened or restored page time to send a heartbeat.
    time.sleep(2.5)

    try:
        with urlopen(
            health_url,
            timeout=0.5,
        ) as response:
            if response.status == 200:
                raw = response.read().decode("utf-8")
                refreshed_health = json.loads(raw or "{}")

                if refreshed_health.get(
                    "browser_active",
                    False,
                ):
                    LOGGER.info(
                        "Browser session became active; "
                        "no duplicate page was opened."
                    )
                    return True

    except (
        OSError,
        URLError,
        json.JSONDecodeError,
    ):
        pass

    LOGGER.info(
        "Server is running without a browser session; "
        "opening %s",
        url,
    )

    open_application_window(url)
    return True
def release_single_instance() -> None:
    global _WINDOWS_MUTEX_HANDLE
    if os.name == "nt" and _WINDOWS_MUTEX_HANDLE is not None:
        ctypes.windll.kernel32.CloseHandle(
            ctypes.c_void_p(_WINDOWS_MUTEX_HANDLE)
        )
    _WINDOWS_MUTEX_HANDLE = None


def report_startup_failure(exc: BaseException) -> None:
    details = (
        f"[{now_iso()}]\n"
        f"{''.join(traceback.format_exception(type(exc), exc, exc.__traceback__))}\n"
    )
    log_written = False
    try:
        STARTUP_ERROR_LOG.write_text(details, encoding="utf-8")
        log_written = True
    except OSError:
        pass
    message = f"تعذّر تشغيل البرنامج:\n{exc}"
    if log_written:
        message += f"\n\nحُفظت التفاصيل في:\n{STARTUP_ERROR_LOG}"
    if os.name == "nt":
        user32 = ctypes.windll.user32
        user32.MessageBoxW.argtypes = [
            ctypes.c_void_p,
            ctypes.c_wchar_p,
            ctypes.c_wchar_p,
            ctypes.c_uint,
        ]
        user32.MessageBoxW.restype = ctypes.c_int
        user32.MessageBoxW(
            None,
            message,
            "تعذّر تشغيل SchemaCraft",
            0x10,
        )
    else:
        print(message, file=sys.stderr)

def run() -> None:
    ensure_storage()

    if not APP_DIR.is_dir():
        raise ApplicationError(
            "مجلد واجهة التطبيق app غير موجود."
        )

    port = application_port()

    server = DataEntryHTTPServer(
        (HOST, port),
        DataEntryRequestHandler,
    )

    url = f"http://{HOST}:{port}/"

    threading.Thread(
        target=open_browser_when_needed,
        args=(server, url),
        daemon=True,
        name="browser-open-check",
    ).start()

    LOGGER.info(
        "Application server started at %s",
        url,
    )

    print(
        f"Data Entry Builder is running at {url}"
    )

    try:
        server.serve_forever()

    except KeyboardInterrupt:
        LOGGER.info(
            "Application interrupted from the console."
        )

    finally:
        server.server_close()

        LOGGER.info(
            "Application server stopped."
        )
if __name__ == "__main__":
    try:
        configure_logging()
        install_exception_logging()

        LOGGER.info(
            "Application process starting. "
            "Frozen=%s Base=%s",
            getattr(sys, "frozen", False),
            BASE_DIR,
        )

        owns_instance = acquire_single_instance()

        if not owns_instance:
            if open_running_instance():
                raise SystemExit(0)

            raise ApplicationError(
                "البرنامج مفتوح بالفعل، "
                "لكن تعذّر الاتصال بنافذته الحالية."
            )

        run()

    except SystemExit:
        raise

    except Exception as error:
        LOGGER.exception(
            "Application startup or runtime failure"
        )

        report_startup_failure(error)
        raise SystemExit(1) from error

    finally:
        release_single_instance()

        if LOGGER.handlers:
            LOGGER.info(
                "Application process finished."
            )
